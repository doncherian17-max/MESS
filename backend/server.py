from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env", override=False)

import os
import io
import logging
import secrets
from datetime import datetime, timedelta, date as date_cls, timezone as _tz
from typing import List, Optional, Literal

from zoneinfo import ZoneInfo

import bcrypt
import jwt
import httpx
import csv as csvlib
from bson import ObjectId
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, Request, BackgroundTasks, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------- Config ----------------
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALG = "HS256"
ACCESS_TOKEN_MIN = 60 * 24 * 7
TZ = ZoneInfo(os.environ.get("TIMEZONE", "Asia/Kolkata"))
APP_URL = os.environ.get("APP_URL", "").rstrip("/")

BREAKFAST_CUTOFF_HOUR = 23
BREAKFAST_CUTOFF_MIN = 30
DINNER_CUTOFF_HOUR = 15
DINNER_CUTOFF_MIN = 0

# Breakfast booking window opens at 10:00 on the day before the meal
BREAKFAST_OPEN_HOUR = 10
BREAKFAST_OPEN_MIN = 0

MAX_QTY = 5

# Email
EMAIL_BASE_URL = "https://integrations.emergentagent.com"
EMAIL_KEY = os.environ.get("EMERGENT_EMAIL_KEY", "")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "MessBook")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="MessBook API")
api = APIRouter(prefix="/api")
bearer = HTTPBearer(auto_error=False)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("mess")

Role = Literal["employee", "admin", "chef"]


# ---------------- Helpers ----------------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_token(user_id: str, employee_number: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "emp": employee_number,
        "role": role,
        "exp": datetime.now(tz=TZ) + timedelta(minutes=ACCESS_TOKEN_MIN),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def now_local() -> datetime:
    return datetime.now(tz=TZ)


def parse_iso_date(s: str) -> date_cls:
    return date_cls.fromisoformat(s)


def user_public(u: dict) -> dict:
    return {
        "id": str(u["_id"]),
        "employee_number": u["employee_number"],
        "name": u.get("name", ""),
        "email": u.get("email"),
        "role": u.get("role", "employee"),
        "created_at": u.get("created_at"),
    }


DINNER_OPEN_HOUR = 19  # Next-day dinner window opens 7 PM the previous day
DINNER_OPEN_MIN = 0


def compute_cutoff(meal_type: str, meal_date: date_cls) -> datetime:
    if meal_type == "breakfast":
        cutoff_day = meal_date - timedelta(days=1)
        return datetime(cutoff_day.year, cutoff_day.month, cutoff_day.day,
                        BREAKFAST_CUTOFF_HOUR, BREAKFAST_CUTOFF_MIN, tzinfo=TZ)
    return datetime(meal_date.year, meal_date.month, meal_date.day,
                    DINNER_CUTOFF_HOUR, DINNER_CUTOFF_MIN, tzinfo=TZ)


def compute_opens_at(meal_type: str, meal_date: date_cls) -> Optional[datetime]:
    """When the booking window opens.

    Breakfast: 10:00 AM the day before.
    Dinner: 7:00 PM the day before (Next-Day Dinner window).
    """
    if meal_type == "breakfast":
        open_day = meal_date - timedelta(days=1)
        return datetime(open_day.year, open_day.month, open_day.day,
                        BREAKFAST_OPEN_HOUR, BREAKFAST_OPEN_MIN, tzinfo=TZ)
    open_day = meal_date - timedelta(days=1)
    return datetime(open_day.year, open_day.month, open_day.day,
                    DINNER_OPEN_HOUR, DINNER_OPEN_MIN, tzinfo=TZ)


async def is_sunday_blocked(meal_date: str, meal_type: str) -> bool:
    """Sundays are Mess Off by default per meal type.

    Admin can whitelist a specific Sunday for `breakfast`, `dinner`, or `both`
    via the `sunday_overrides` collection. Returns True only if the date is a
    Sunday AND no override covers the requested meal_type.
    """
    d = parse_iso_date(meal_date)
    if d.weekday() != 6:  # 6 = Sunday
        return False
    ov = await db.sunday_overrides.find_one({"date": meal_date})
    if not ov:
        return True
    meals = ov.get("meals", "both")
    return meals != "both" and meals != meal_type


async def audit(actor: dict, action: str, target: str = "", meta: Optional[dict] = None):
    try:
        await db.audit_logs.insert_one({
            "actor_id": str(actor.get("_id", "")) if actor else "",
            "actor_employee_number": actor.get("employee_number", "") if actor else "",
            "actor_role": actor.get("role", "") if actor else "",
            "action": action,
            "target": target,
            "meta": meta or {},
            "timestamp": now_local().isoformat(),
        })
    except Exception as e:
        logger.warning(f"audit log failed: {e}")


async def record_cancellation(booking: dict, cancelled_by: str, actor_role: str, reason: str = ""):
    """Log a cancellation event so employees can see which of their meals were cancelled and why."""
    try:
        await db.cancellation_events.insert_one({
            "user_id": booking["user_id"],
            "employee_number": booking.get("employee_number", ""),
            "employee_name": booking.get("employee_name", ""),
            "meal_type": booking["meal_type"],
            "meal_date": booking["meal_date"],
            "quantity": booking.get("quantity", 1),
            "booking_type": booking.get("booking_type", "dine_in"),
            "cancelled_by": cancelled_by,       # employee_number or "system"
            "actor_role": actor_role,           # "employee" | "admin" | "system"
            "reason": reason,
            "cancelled_at": now_local().isoformat(),
        })
    except Exception as e:
        logger.warning(f"cancellation log failed: {e}")


async def send_apology_email(user_doc: dict, meal_date: str, meals: list, reason: str, background: BackgroundTasks):
    """Email dispatch disabled — retained as a no-op for backwards compatibility."""
    return False


async def send_email_async(to: str, subject: str, html: str) -> bool:
    """Email dispatch disabled — retained as a no-op for backwards compatibility."""
    return False


# ---------------- Models ----------------
class LoginIn(BaseModel):
    employee_number: str
    password: str


class AdminCreateEmployeeIn(BaseModel):
    employee_number: str = Field(min_length=1, max_length=32)
    name: str = Field(min_length=1, max_length=100)
    password: str = Field(min_length=4, max_length=128)
    role: Literal["employee", "admin", "chef"] = "employee"


class BookingIn(BaseModel):
    meal_type: Literal["breakfast", "dinner"]
    meal_date: str
    quantity: int = Field(default=1, ge=1, le=MAX_QTY)
    booking_type: Literal["dine_in", "parcel"] = "dine_in"


class BookingUpdateIn(BaseModel):
    quantity: Optional[int] = Field(default=None, ge=1, le=MAX_QTY)
    booking_type: Optional[Literal["dine_in", "parcel"]] = None


class UpdateMeIn(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=100)


class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=4, max_length=128)


class HolidayIn(BaseModel):
    date: str
    end_date: Optional[str] = None  # inclusive range end; if None, single day
    name: str = Field(min_length=1, max_length=100)
    applies_to: Literal["breakfast", "dinner", "both"] = "both"


class MealPricesIn(BaseModel):
    breakfast: float = Field(ge=0)
    dinner: float = Field(ge=0)


class DeleteBookingsRangeIn(BaseModel):
    from_date: str
    to_date: str
    meal_type: Optional[Literal["breakfast", "dinner"]] = None


class MenuIn(BaseModel):
    date: str
    meal_type: Literal["breakfast", "dinner"]
    items: List[str] = Field(default_factory=list)


WEEKDAY_NAMES = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]


class WeeklyMenuIn(BaseModel):
    day_of_week: Literal["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    meal_type: Literal["breakfast", "dinner"]
    items: List[str] = Field(default_factory=list)


async def get_current_user(
    request: Request,
    creds: Optional[HTTPAuthorizationCredentials] = Depends(bearer),
) -> dict:
    token = None
    if creds and creds.scheme.lower() == "bearer":
        token = creds.credentials
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    try:
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


async def get_admin_user(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


async def get_chef_or_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") not in ("chef", "admin"):
        raise HTTPException(status_code=403, detail="Chef/Admin access required")
    return user


# ---------------- Auth endpoints ----------------
@api.post("/auth/login")
async def login(body: LoginIn):
    emp = body.employee_number.strip()
    user = await db.users.find_one({"employee_number": emp})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid employee number or password")
    token = create_token(str(user["_id"]), emp, user.get("role", "employee"))
    return {"token": token, "user": user_public(user)}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user_public(user)


@api.patch("/auth/me")
async def update_me(body: UpdateMeIn, user: dict = Depends(get_current_user)):
    updates = {}
    if body.name is not None:
        updates["name"] = body.name.strip()
    if updates:
        await db.users.update_one({"_id": user["_id"]}, {"$set": updates})
        user.update(updates)
    return user_public(user)


@api.post("/auth/change-password")
async def change_password(body: ChangePasswordIn, user: dict = Depends(get_current_user)):
    if not verify_password(body.current_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"password_hash": hash_password(body.new_password)}})
    await audit(user, "user.change_password")
    return {"ok": True}


# ---------------- Meal prices (settings) ----------------
DEFAULT_BREAKFAST_PRICE = 50.0
DEFAULT_DINNER_PRICE = 80.0


async def get_meal_prices() -> dict:
    doc = await db.settings.find_one({"key": "meal_prices"})
    if doc:
        return {
            "breakfast": float(doc.get("breakfast", DEFAULT_BREAKFAST_PRICE)),
            "dinner": float(doc.get("dinner", DEFAULT_DINNER_PRICE)),
        }
    return {"breakfast": DEFAULT_BREAKFAST_PRICE, "dinner": DEFAULT_DINNER_PRICE}


@api.get("/settings/prices")
async def get_prices(user: dict = Depends(get_current_user)):
    return await get_meal_prices()


@api.put("/admin/settings/prices")
async def set_prices(body: MealPricesIn, admin: dict = Depends(get_admin_user)):
    await db.settings.update_one(
        {"key": "meal_prices"},
        {"$set": {"breakfast": float(body.breakfast), "dinner": float(body.dinner),
                  "updated_at": now_local().isoformat()}},
        upsert=True,
    )
    await audit(admin, "settings.prices.update",
                meta={"breakfast": body.breakfast, "dinner": body.dinner})
    return {"ok": True, "breakfast": float(body.breakfast), "dinner": float(body.dinner)}


# ---------------- Bookings ----------------
async def is_emergency_cancelled(meal_date: str, meal_type: str, user_id: Optional[str] = None) -> Optional[dict]:
    """Return the active emergency-cancellation record blocking this meal, if any."""
    query = {
        "date": meal_date,
        "meal_type": {"$in": [meal_type, "both"]},
        "active": True,
    }
    async for ec in db.emergency_cancellations.find(query):
        if ec.get("applies_to") == "all":
            return ec
        if user_id and user_id in (ec.get("employee_ids") or []):
            return ec
    return None


async def is_holiday(meal_date: str, meal_type: str) -> Optional[dict]:
    """Match single-date holidays OR any range where date <= meal_date <= end_date."""
    # Single-day match
    h = await db.holidays.find_one({
        "date": meal_date,
        "applies_to": {"$in": [meal_type, "both"]},
    })
    if h:
        return h
    # Range match: date <= meal_date AND (end_date exists AND end_date >= meal_date)
    async for cand in db.holidays.find({
        "date": {"$lte": meal_date},
        "end_date": {"$gte": meal_date},
        "applies_to": {"$in": [meal_type, "both"]},
    }):
        return cand
    return None


@api.get("/bookings/status")
async def booking_status(user: dict = Depends(get_current_user)):
    now = now_local()
    today = now.date()
    tomorrow = today + timedelta(days=1)

    # Next-Day Dinner rolling window: after today's 3 PM cutoff, the dinner card rolls to TOMORROW.
    today_dinner_cutoff = compute_cutoff("dinner", today)
    dinner_target = today if now < today_dinner_cutoff else tomorrow

    results = []
    for meal_type, meal_date in [("breakfast", tomorrow), ("dinner", dinner_target)]:
        cutoff = compute_cutoff(meal_type, meal_date)
        opens_at = compute_opens_at(meal_type, meal_date)
        existing = await db.bookings.find_one({
            "user_id": str(user["_id"]),
            "meal_type": meal_type,
            "meal_date": meal_date.isoformat(),
        })
        holiday = await is_holiday(meal_date.isoformat(), meal_type)
        sunday_off = await is_sunday_blocked(meal_date.isoformat(), meal_type)
        ec = await is_emergency_cancelled(meal_date.isoformat(), meal_type, str(user["_id"]))
        cancellation = None
        if ec:
            cancellation = {
                "reason": ec.get("reason", ""),
                "meal_type": ec.get("meal_type", meal_type),
                "date": ec.get("date", meal_date.isoformat()),
                "created_at": ec.get("created_at"),
            }
        day_label = "Today" if meal_date == today else ("Tomorrow" if meal_date == tomorrow else meal_date.isoformat())
        results.append({
            "meal_type": meal_type,
            "meal_date": meal_date.isoformat(),
            "day_label": day_label,
            "cutoff": cutoff.isoformat(),
            "cutoff_passed": now >= cutoff,
            "opens_at": opens_at.isoformat() if opens_at else None,
            "not_yet_open": bool(opens_at and now < opens_at),
            "booked": bool(existing),
            "booking_id": str(existing["_id"]) if existing else None,
            "quantity": existing.get("quantity", 1) if existing else None,
            "booking_type": existing.get("booking_type", "dine_in") if existing else None,
            "holiday": {"name": holiday["name"]} if holiday else None,
            "sunday_off": sunday_off,
            "cancellation": cancellation,
        })
    return {"now": now.isoformat(), "items": results}


@api.post("/bookings")
async def create_booking(body: BookingIn, user: dict = Depends(get_current_user)):
    try:
        meal_date = parse_iso_date(body.meal_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid meal_date, use YYYY-MM-DD")

    now = now_local()
    cutoff = compute_cutoff(body.meal_type, meal_date)
    opens_at = compute_opens_at(body.meal_type, meal_date)
    if opens_at and now < opens_at:
        raise HTTPException(status_code=400, detail=f"Booking opens at {opens_at.strftime('%I:%M %p on %d %b')}")
    if now >= cutoff:
        raise HTTPException(status_code=400, detail=f"Booking cutoff has passed (cutoff was {cutoff.strftime('%d %b %Y %I:%M %p')})")
    if meal_date < now.date() - timedelta(days=1):
        raise HTTPException(status_code=400, detail="Cannot book for past dates")

    holiday = await is_holiday(body.meal_date, body.meal_type)
    if holiday:
        raise HTTPException(status_code=400, detail=f"{body.meal_type.capitalize()} not available: {holiday['name']} holiday")

    if await is_sunday_blocked(body.meal_date, body.meal_type):
        raise HTTPException(status_code=400, detail=f"Sunday is Mess Off — {body.meal_type} is not available. Please contact the admin if bookings should be opened.")

    emerg = await is_emergency_cancelled(body.meal_date, body.meal_type, user_id=str(user["_id"]))
    if emerg:
        raise HTTPException(status_code=400, detail=f"Bookings for this {body.meal_type} are closed by admin: {emerg['reason']}")

    existing = await db.bookings.find_one({
        "user_id": str(user["_id"]),
        "meal_type": body.meal_type,
        "meal_date": body.meal_date,
    })
    if existing:
        if existing.get("status") == "emergency_cancelled":
            # Revive the soft-cancelled row instead of erroring — happens after admin reopens the emergency
            await db.bookings.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "status": "active",
                    "quantity": body.quantity,
                    "booking_type": body.booking_type,
                    "created_at": now.isoformat(),
                }, "$unset": {"cancelled_at": "", "cancelled_by": ""}},
            )
            await audit(user, "booking.rebook_after_reopen", target=str(existing["_id"]),
                        meta={"meal_type": body.meal_type, "meal_date": body.meal_date,
                              "qty": body.quantity, "type": body.booking_type})
            return {"id": str(existing["_id"]), "meal_type": body.meal_type, "meal_date": body.meal_date,
                    "quantity": body.quantity, "booking_type": body.booking_type}
        raise HTTPException(status_code=400, detail="You have already booked this meal. You can update or cancel it.")

    doc = {
        "user_id": str(user["_id"]),
        "employee_number": user["employee_number"],
        "employee_name": user.get("name", ""),
        "meal_type": body.meal_type,
        "meal_date": body.meal_date,
        "quantity": body.quantity,
        "booking_type": body.booking_type,
        "served": False,
        "served_at": None,
        "served_by": None,
        "created_at": now.isoformat(),
    }
    res = await db.bookings.insert_one(doc)
    await audit(user, "booking.create", target=str(res.inserted_id), meta={
        "meal_type": body.meal_type, "meal_date": body.meal_date, "qty": body.quantity, "type": body.booking_type,
    })
    return {"id": str(res.inserted_id), "meal_type": body.meal_type, "meal_date": body.meal_date,
            "quantity": body.quantity, "booking_type": body.booking_type}


@api.patch("/bookings/{booking_id}")
async def update_booking(booking_id: str, body: BookingUpdateIn, user: dict = Depends(get_current_user)):
    try:
        booking = await db.bookings.find_one({"_id": ObjectId(booking_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid booking id")
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking["user_id"] != str(user["_id"]):
        raise HTTPException(status_code=403, detail="Not your booking")
    if booking.get("served"):
        raise HTTPException(status_code=400, detail="Booking already served, cannot update")

    meal_date = parse_iso_date(booking["meal_date"])
    cutoff = compute_cutoff(booking["meal_type"], meal_date)
    if now_local() >= cutoff:
        raise HTTPException(status_code=400, detail="Cannot update after cutoff")

    updates = {}
    if body.quantity is not None:
        updates["quantity"] = body.quantity
    if body.booking_type is not None:
        updates["booking_type"] = body.booking_type
    if updates:
        await db.bookings.update_one({"_id": ObjectId(booking_id)}, {"$set": updates})
        await audit(user, "booking.update", target=booking_id, meta=updates)
    return {"ok": True, **updates}


@api.delete("/bookings/{booking_id}")
async def cancel_booking(booking_id: str, user: dict = Depends(get_current_user)):
    try:
        booking = await db.bookings.find_one({"_id": ObjectId(booking_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid booking id")
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking["user_id"] != str(user["_id"]):
        raise HTTPException(status_code=403, detail="Not your booking")
    if booking.get("served"):
        raise HTTPException(status_code=400, detail="Booking already served, cannot cancel")

    meal_date = parse_iso_date(booking["meal_date"])
    cutoff = compute_cutoff(booking["meal_type"], meal_date)
    if now_local() >= cutoff:
        raise HTTPException(status_code=400, detail="Cannot cancel after cutoff")

    await db.bookings.delete_one({"_id": ObjectId(booking_id)})
    await record_cancellation(booking, cancelled_by=user["employee_number"],
                              actor_role="employee", reason="Cancelled by employee")
    await audit(user, "booking.cancel", target=booking_id)
    return {"ok": True}


@api.get("/bookings/cancellations")
async def my_cancellations(
    month: Optional[str] = Query(None, description="YYYY-MM to filter"),
    user: dict = Depends(get_current_user),
):
    """List cancellation events for the current user (defaults to current month)."""
    if not month:
        n = now_local()
        month = f"{n.year:04d}-{n.month:02d}"
    items = []
    cursor = db.cancellation_events.find({
        "user_id": str(user["_id"]),
        "meal_date": {"$regex": f"^{month}-"},
    }).sort("cancelled_at", -1)
    async for e in cursor:
        items.append({
            "id": str(e["_id"]),
            "meal_type": e["meal_type"],
            "meal_date": e["meal_date"],
            "quantity": e.get("quantity", 1),
            "booking_type": e.get("booking_type", "dine_in"),
            "cancelled_by": e.get("cancelled_by", ""),
            "actor_role": e.get("actor_role", ""),
            "reason": e.get("reason", ""),
            "cancelled_at": e.get("cancelled_at"),
        })
    return {"month": month, "items": items}


@api.get("/bookings/mine")
async def my_bookings(
    user: dict = Depends(get_current_user),
):
    """Return only the CURRENT MONTH's bookings for the logged-in employee."""
    n = now_local()
    month = f"{n.year:04d}-{n.month:02d}"
    cursor = db.bookings.find({
        "user_id": str(user["_id"]),
        "meal_date": {"$regex": f"^{month}-"},
        "status": {"$ne": "emergency_cancelled"},
    }).sort("meal_date", 1)
    items = []
    b_qty = d_qty = 0
    async for b in cursor:
        q = b.get("quantity", 1)
        if b["meal_type"] == "breakfast":
            b_qty += q
        else:
            d_qty += q
        items.append({
            "id": str(b["_id"]),
            "meal_type": b["meal_type"],
            "meal_date": b["meal_date"],
            "quantity": q,
            "booking_type": b.get("booking_type", "dine_in"),
            "served": b.get("served", False),
            "created_at": b.get("created_at"),
        })
    # Deduction totals in ₹
    prices = await get_meal_prices()
    deduction = round(b_qty * prices["breakfast"] + d_qty * prices["dinner"], 2)
    return {
        "month": month,
        "breakfast_count": b_qty,
        "dinner_count": d_qty,
        "total": b_qty + d_qty,
        "breakfast_price": prices["breakfast"],
        "dinner_price": prices["dinner"],
        "deduction": deduction,
        "items": items,
    }


@api.get("/health")
async def health():
    return {"status": "ok"}


# ---------------- Menu (read for all authed users) ----------------
def _weekday_key(date_str: str) -> str:
    d = parse_iso_date(date_str)
    return WEEKDAY_NAMES[d.weekday()]


async def _resolve_menu_for(date_str: str, meal_type: str) -> List[str]:
    """Resolve menu items for a date/meal. Date-specific overrides weekly template."""
    doc = await db.menus.find_one({"date": date_str, "meal_type": meal_type})
    if doc and doc.get("items"):
        return doc.get("items", [])
    dow = _weekday_key(date_str)
    wdoc = await db.weekly_menus.find_one({"day_of_week": dow, "meal_type": meal_type})
    if wdoc:
        return wdoc.get("items", [])
    return []


@api.get("/menu")
async def get_menu(date: str, meal_type: Optional[Literal["breakfast", "dinner"]] = None,
                   user: dict = Depends(get_current_user)):
    meals = [meal_type] if meal_type else ["breakfast", "dinner"]
    items = []
    for mt in meals:
        resolved = await _resolve_menu_for(date, mt)
        items.append({"date": date, "meal_type": mt, "items": resolved})
    return items


@api.get("/weekly-menu")
async def get_weekly_menu(user: dict = Depends(get_current_user)):
    """Return the full weekly menu template (7 days x 2 meals)."""
    out = {}
    async for m in db.weekly_menus.find({}):
        out[f"{m['day_of_week']}:{m['meal_type']}"] = m.get("items", [])
    result = []
    for dow in WEEKDAY_NAMES:
        for mt in ("breakfast", "dinner"):
            result.append({
                "day_of_week": dow,
                "meal_type": mt,
                "items": out.get(f"{dow}:{mt}", []),
            })
    return result


@api.get("/weekly-menu/today")
async def get_weekly_menu_today(user: dict = Depends(get_current_user)):
    """Return today's resolved menu (date-specific overrides weekly)."""
    today = now_local().date().isoformat()
    breakfast = await _resolve_menu_for(today, "breakfast")
    dinner = await _resolve_menu_for(today, "dinner")
    return {
        "date": today,
        "day_of_week": _weekday_key(today),
        "breakfast": breakfast,
        "dinner": dinner,
    }


@api.get("/holidays")
async def list_holidays(user: dict = Depends(get_current_user)):
    now = now_local().date()
    start = (now - timedelta(days=7)).isoformat()
    end = (now + timedelta(days=60)).isoformat()
    out = []
    async for h in db.holidays.find({"date": {"$gte": start, "$lte": end}}).sort("date", 1):
        out.append({"id": str(h["_id"]), "date": h["date"], "name": h["name"], "applies_to": h.get("applies_to", "both")})
    return out


# ---------------- Chef endpoints ----------------
@api.get("/chef/summary")
async def chef_summary(date: Optional[str] = None, u: dict = Depends(get_chef_or_admin)):
    d = date or now_local().date().isoformat()
    out = {"date": d, "breakfast": {}, "dinner": {}}
    for meal in ("breakfast", "dinner"):
        pipeline = [
            {"$match": {"meal_date": d, "meal_type": meal, "status": {"$ne": "emergency_cancelled"}}},
            {"$group": {
                "_id": {"type": "$booking_type", "served": "$served"},
                "qty": {"$sum": "$quantity"},
                "orders": {"$sum": 1},
            }},
        ]
        parcel_qty = dine_qty = served_qty = pending_qty = total_qty = total_orders = 0
        async for row in db.bookings.aggregate(pipeline):
            btype = row["_id"].get("type", "dine_in")
            served = row["_id"].get("served", False)
            q = row["qty"]
            total_qty += q
            total_orders += row["orders"]
            if btype == "parcel":
                parcel_qty += q
            else:
                dine_qty += q
            if served:
                served_qty += q
            else:
                pending_qty += q
        out[meal] = {
            "total": total_qty,
            "orders": total_orders,
            "parcel": parcel_qty,
            "dine_in": dine_qty,
            "served": served_qty,
            "pending": pending_qty,
        }
    return out


@api.get("/chef/bookings")
async def chef_bookings(
    date: Optional[str] = None,
    meal_type: Optional[Literal["breakfast", "dinner"]] = None,
    q: Optional[str] = None,
    u: dict = Depends(get_chef_or_admin),
):
    d = date or now_local().date().isoformat()
    query: dict = {"meal_date": d, "status": {"$ne": "emergency_cancelled"}}
    if meal_type:
        query["meal_type"] = meal_type
    if q:
        query["$or"] = [
            {"employee_number": {"$regex": q, "$options": "i"}},
            {"employee_name": {"$regex": q, "$options": "i"}},
        ]
    items = []
    async for b in db.bookings.find(query).sort([("meal_type", 1), ("employee_number", 1)]):
        items.append({
            "id": str(b["_id"]),
            "employee_number": b["employee_number"],
            "employee_name": b.get("employee_name", ""),
            "meal_type": b["meal_type"],
            "meal_date": b["meal_date"],
            "quantity": b.get("quantity", 1),
            "booking_type": b.get("booking_type", "dine_in"),
            "served": b.get("served", False),
            "served_at": b.get("served_at"),
            "served_by": b.get("served_by"),
            "admin_override": b.get("admin_override", False),
            "override_reason": b.get("override_reason"),
        })
    return items


@api.post("/chef/serve/{booking_id}")
async def chef_serve(booking_id: str, u: dict = Depends(get_chef_or_admin)):
    try:
        oid = ObjectId(booking_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    b = await db.bookings.find_one({"_id": oid})
    if not b:
        raise HTTPException(status_code=404, detail="Not found")
    if b.get("served"):
        raise HTTPException(status_code=400, detail="Already served")
    await db.bookings.update_one({"_id": oid}, {"$set": {
        "served": True,
        "served_at": now_local().isoformat(),
        "served_by": u["employee_number"],
    }})
    await audit(u, "booking.serve", target=booking_id)
    return {"ok": True}


@api.post("/chef/unserve/{booking_id}")
async def chef_unserve(booking_id: str, u: dict = Depends(get_chef_or_admin)):
    try:
        oid = ObjectId(booking_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    await db.bookings.update_one({"_id": oid}, {"$set": {
        "served": False, "served_at": None, "served_by": None,
    }})
    await audit(u, "booking.unserve", target=booking_id)
    return {"ok": True}


# ---------------- Admin endpoints ----------------
@api.get("/admin/employees")
async def list_employees(admin: dict = Depends(get_admin_user)):
    users = []
    async for u in db.users.find({}).sort("created_at", -1):
        users.append(user_public(u))
    return users


@api.post("/admin/employees")
async def admin_create_employee(body: AdminCreateEmployeeIn, admin: dict = Depends(get_admin_user)):
    emp = body.employee_number.strip()
    if await db.users.find_one({"employee_number": emp}):
        raise HTTPException(status_code=400, detail="Employee number already exists")
    doc = {
        "employee_number": emp,
        "name": body.name.strip(),
        "email": None,
        "password_hash": hash_password(body.password),
        "role": body.role,
        "created_at": now_local().isoformat(),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    await audit(admin, "employee.create", target=emp, meta={"role": body.role})
    return user_public(doc)


class AdminUpdateEmployeeIn(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=100)
    role: Optional[Literal["employee", "admin", "chef"]] = None
    password: Optional[str] = Field(default=None, min_length=4, max_length=128)


@api.patch("/admin/employees/{user_id}")
async def admin_update_employee(user_id: str, body: AdminUpdateEmployeeIn, admin: dict = Depends(get_admin_user)):
    try:
        oid = ObjectId(user_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    target = await db.users.find_one({"_id": oid})
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    updates: dict = {}
    audit_meta: dict = {}
    if body.name is not None:
        updates["name"] = body.name.strip()
        audit_meta["name"] = updates["name"]
    if body.role is not None:
        if str(oid) == str(admin["_id"]) and body.role != "admin":
            raise HTTPException(status_code=400, detail="Cannot change your own role")
        updates["role"] = body.role
        audit_meta["role"] = body.role
    if body.password is not None:
        updates["password_hash"] = hash_password(body.password)
        audit_meta["password_reset"] = True
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    await db.users.update_one({"_id": oid}, {"$set": updates})
    await audit(admin, "employee.update", target=target.get("employee_number", ""), meta=audit_meta)
    target.update({k: v for k, v in updates.items() if k != "password_hash"})
    return user_public(target)


@api.delete("/admin/employees/{user_id}")
async def admin_delete_employee(user_id: str, admin: dict = Depends(get_admin_user)):
    try:
        oid = ObjectId(user_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    if str(oid) == str(admin["_id"]):
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    target = await db.users.find_one({"_id": oid})
    res = await db.users.delete_one({"_id": oid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    await db.bookings.delete_many({"user_id": str(oid)})
    await audit(admin, "employee.delete", target=target.get("employee_number", "") if target else "")
    return {"ok": True}


# --- Admin deductions (monthly meal totals in ₹) ---
@api.get("/admin/deductions")
async def admin_deductions(month: Optional[str] = Query(None), admin: dict = Depends(get_admin_user)):
    """Return per-employee current-month meal counts and ₹ deductions.
    `month` optional in YYYY-MM (defaults to current month)."""
    if not month:
        n = now_local()
        month = f"{n.year:04d}-{n.month:02d}"
    try:
        parse_iso_date(f"{month}-01")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid month, use YYYY-MM")

    prices = await get_meal_prices()
    # Aggregate active bookings per user for this month
    pipeline = [
        {"$match": {
            "meal_date": {"$regex": f"^{month}-"},
            "status": {"$ne": "emergency_cancelled"},
        }},
        {"$group": {
            "_id": {"user_id": "$user_id", "meal_type": "$meal_type"},
            "qty": {"$sum": "$quantity"},
        }},
    ]
    counts: dict = {}  # user_id -> {"breakfast": n, "dinner": n}
    async for row in db.bookings.aggregate(pipeline):
        uid = row["_id"]["user_id"]
        counts.setdefault(uid, {"breakfast": 0, "dinner": 0})
        counts[uid][row["_id"]["meal_type"]] = row["qty"]

    # Merge with user records
    result = []
    async for u in db.users.find({}).sort("employee_number", 1):
        uid = str(u["_id"])
        c = counts.get(uid, {"breakfast": 0, "dinner": 0})
        b_amt = round(c["breakfast"] * prices["breakfast"], 2)
        d_amt = round(c["dinner"] * prices["dinner"], 2)
        result.append({
            "user_id": uid,
            "employee_number": u.get("employee_number", ""),
            "name": u.get("name", ""),
            "role": u.get("role", "employee"),
            "breakfast_count": c["breakfast"],
            "dinner_count": c["dinner"],
            "breakfast_amount": b_amt,
            "dinner_amount": d_amt,
            "total_amount": round(b_amt + d_amt, 2),
        })
    return {"month": month, "prices": prices, "employees": result}


# --- Admin: delete bookings in date range ---
@api.post("/admin/bookings/range-delete")
async def admin_delete_bookings_range(body: DeleteBookingsRangeIn, admin: dict = Depends(get_admin_user)):
    try:
        parse_iso_date(body.from_date); parse_iso_date(body.to_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date")
    if body.to_date < body.from_date:
        raise HTTPException(status_code=400, detail="'to' date must be on or after 'from' date")
    query: dict = {"meal_date": {"$gte": body.from_date, "$lte": body.to_date}}
    if body.meal_type:
        query["meal_type"] = body.meal_type
    res = await db.bookings.delete_many(query)
    await audit(admin, "admin.bookings.range_delete",
                target=f"{body.from_date}..{body.to_date}",
                meta={"deleted": res.deleted_count, "meal_type": body.meal_type or "both"})
    return {"ok": True, "deleted": res.deleted_count}


# --- Holidays ---
@api.get("/admin/holidays")
async def admin_list_holidays(admin: dict = Depends(get_admin_user)):
    out = []
    async for h in db.holidays.find({}).sort("date", -1):
        out.append({
            "id": str(h["_id"]),
            "date": h["date"],
            "end_date": h.get("end_date", h["date"]),
            "name": h["name"],
            "applies_to": h.get("applies_to", "both"),
        })
    return out


@api.post("/admin/holidays")
async def admin_add_holiday(body: HolidayIn, admin: dict = Depends(get_admin_user)):
    try:
        parse_iso_date(body.date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date")
    end_date = body.end_date
    if end_date:
        try:
            if parse_iso_date(end_date) < parse_iso_date(body.date):
                raise HTTPException(status_code=400, detail="End date must be on or after start date")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid end date")
    else:
        end_date = body.date
    # Overlap check: reject if any existing holiday overlaps
    overlap = await db.holidays.find_one({
        "$or": [
            {"date": {"$lte": end_date}, "end_date": {"$gte": body.date}},
        ]
    })
    if overlap:
        raise HTTPException(status_code=400, detail="A holiday already exists in this range")
    doc = {"date": body.date, "end_date": end_date, "name": body.name.strip(),
           "applies_to": body.applies_to, "created_at": now_local().isoformat()}
    res = await db.holidays.insert_one(doc)
    await audit(admin, "holiday.create", target=body.date,
                meta={"name": body.name, "end_date": end_date})
    return {"id": str(res.inserted_id), "date": body.date, "end_date": end_date,
            "name": body.name, "applies_to": body.applies_to}


@api.delete("/admin/holidays/{holiday_id}")
async def admin_delete_holiday(holiday_id: str, admin: dict = Depends(get_admin_user)):
    try:
        oid = ObjectId(holiday_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    await db.holidays.delete_one({"_id": oid})
    await audit(admin, "holiday.delete", target=holiday_id)
    return {"ok": True}


# --- Sunday overrides (open bookings for a specific Sunday) ---
class SundayOverrideIn(BaseModel):
    date: str
    meals: Literal["breakfast", "dinner", "both"] = "both"


@api.get("/admin/sunday-overrides")
async def admin_list_sunday_overrides(admin: dict = Depends(get_admin_user)):
    out = []
    async for s in db.sunday_overrides.find({}).sort("date", -1):
        out.append({"id": str(s["_id"]), "date": s["date"],
                    "meals": s.get("meals", "both"),
                    "created_by": s.get("created_by", ""),
                    "created_at": s.get("created_at", "")})
    return out


@api.post("/admin/sunday-overrides")
async def admin_add_sunday_override(body: SundayOverrideIn, admin: dict = Depends(get_admin_user)):
    try:
        d = parse_iso_date(body.date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date")
    if d.weekday() != 6:
        raise HTTPException(status_code=400, detail="Sunday overrides only apply to Sundays")
    doc = {"date": body.date, "meals": body.meals,
           "created_by": admin["employee_number"],
           "created_at": now_local().isoformat()}
    await db.sunday_overrides.update_one(
        {"date": body.date},
        {"$set": doc},
        upsert=True,
    )
    await audit(admin, "sunday_override.upsert",
                target=body.date, meta={"meals": body.meals})
    return {"date": body.date, "meals": body.meals}


@api.delete("/admin/sunday-overrides/{date}")
async def admin_delete_sunday_override(date: str, admin: dict = Depends(get_admin_user)):
    res = await db.sunday_overrides.delete_one({"date": date})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No override found for this date")
    await audit(admin, "sunday_override.remove", target=date)
    return {"ok": True}


@api.get("/sunday-off-info")
async def get_sunday_off_info(user: dict = Depends(get_current_user)):
    """Public: list Sundays that admin has explicitly opened (and for which meals)."""
    now = now_local().date()
    out = []
    async for s in db.sunday_overrides.find({"date": {"$gte": now.isoformat()}}).sort("date", 1):
        out.append({"date": s["date"], "meals": s.get("meals", "both")})
    return {"policy": "Sundays are Mess Off by default", "open_sundays": out}


# --- Menu ---
@api.get("/admin/menu")
async def admin_list_menu(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    admin: dict = Depends(get_admin_user),
):
    now = now_local().date()
    start = from_date or (now - timedelta(days=1)).isoformat()
    end = to_date or (now + timedelta(days=14)).isoformat()
    out = []
    async for m in db.menus.find({"date": {"$gte": start, "$lte": end}}).sort([("date", 1), ("meal_type", 1)]):
        out.append({"id": str(m["_id"]), "date": m["date"], "meal_type": m["meal_type"], "items": m.get("items", [])})
    return out


@api.put("/admin/menu")
async def admin_upsert_menu(body: MenuIn, admin: dict = Depends(get_admin_user)):
    try:
        parse_iso_date(body.date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date")
    items = [i.strip() for i in body.items if i.strip()]
    await db.menus.update_one(
        {"date": body.date, "meal_type": body.meal_type},
        {"$set": {"items": items, "updated_at": now_local().isoformat()}},
        upsert=True,
    )
    await audit(admin, "menu.update", target=f"{body.date}:{body.meal_type}", meta={"items": items})
    return {"ok": True, "date": body.date, "meal_type": body.meal_type, "items": items}


@api.delete("/admin/menu/{menu_id}")
async def admin_delete_menu(menu_id: str, admin: dict = Depends(get_admin_user)):
    try:
        oid = ObjectId(menu_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    existing = await db.menus.find_one({"_id": oid})
    if not existing:
        raise HTTPException(status_code=404, detail="Menu item not found")
    await db.menus.delete_one({"_id": oid})
    await audit(admin, "menu.delete", target=f"{existing['date']}:{existing['meal_type']}",
                meta={"date": existing["date"], "meal_type": existing["meal_type"],
                      "items": existing.get("items", [])})
    return {"ok": True, "message": "Menu item deleted successfully."}


# --- Weekly Menu (admin CRUD) ---
@api.get("/admin/weekly-menu")
async def admin_list_weekly_menu(admin: dict = Depends(get_admin_user)):
    """Return the weekly template with all 14 slots (fills empty ones)."""
    out = {}
    async for m in db.weekly_menus.find({}):
        out[f"{m['day_of_week']}:{m['meal_type']}"] = {
            "items": m.get("items", []),
            "updated_at": m.get("updated_at"),
        }
    result = []
    for dow in WEEKDAY_NAMES:
        for mt in ("breakfast", "dinner"):
            key = f"{dow}:{mt}"
            entry = out.get(key, {"items": [], "updated_at": None})
            result.append({
                "day_of_week": dow,
                "meal_type": mt,
                "items": entry["items"],
                "updated_at": entry["updated_at"],
            })
    return result


@api.put("/admin/weekly-menu")
async def admin_upsert_weekly_menu(body: WeeklyMenuIn, admin: dict = Depends(get_admin_user)):
    items = [i.strip() for i in body.items if i.strip()]
    await db.weekly_menus.update_one(
        {"day_of_week": body.day_of_week, "meal_type": body.meal_type},
        {"$set": {"items": items, "updated_at": now_local().isoformat()}},
        upsert=True,
    )
    await audit(admin, "weekly_menu.update",
                target=f"{body.day_of_week}:{body.meal_type}",
                meta={"items": items})
    return {"ok": True, "day_of_week": body.day_of_week, "meal_type": body.meal_type, "items": items}


@api.delete("/admin/weekly-menu")
async def admin_delete_weekly_menu(
    day_of_week: Literal["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"],
    meal_type: Literal["breakfast", "dinner"],
    admin: dict = Depends(get_admin_user),
):
    existing = await db.weekly_menus.find_one({"day_of_week": day_of_week, "meal_type": meal_type})
    if not existing:
        raise HTTPException(status_code=404, detail="Weekly menu slot is already empty")
    await db.weekly_menus.delete_one({"_id": existing["_id"]})
    await audit(admin, "weekly_menu.delete",
                target=f"{day_of_week}:{meal_type}",
                meta={"items": existing.get("items", [])})
    return {"ok": True, "message": "Weekly menu cleared for this slot."}


# --- Reports & Audit ---
@api.get("/admin/summary")
async def admin_summary(
    from_date: str = Query(..., alias="from"),
    to_date: str = Query(..., alias="to"),
    admin: dict = Depends(get_admin_user),
):
    try:
        parse_iso_date(from_date)
        parse_iso_date(to_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")

    pipeline = [
        {"$match": {"meal_date": {"$gte": from_date, "$lte": to_date}, "status": {"$ne": "emergency_cancelled"}}},
        {"$group": {
            "_id": {"emp": "$employee_number", "name": "$employee_name",
                    "meal": "$meal_type", "type": "$booking_type"},
            "qty": {"$sum": "$quantity"},
        }},
    ]
    per_emp: dict = {}
    total_b_dine = total_b_parcel = total_d_dine = total_d_parcel = 0
    async for row in db.bookings.aggregate(pipeline):
        emp = row["_id"]["emp"]
        name = row["_id"].get("name") or ""
        meal = row["_id"]["meal"]
        btype = row["_id"].get("type") or "dine_in"
        cnt = row["qty"]
        if emp not in per_emp:
            per_emp[emp] = {
                "employee_number": emp, "name": name,
                "breakfast_dine_in": 0, "breakfast_parcel": 0,
                "dinner_dine_in": 0, "dinner_parcel": 0,
            }
        if meal == "breakfast":
            if btype == "parcel":
                per_emp[emp]["breakfast_parcel"] += cnt
                total_b_parcel += cnt
            else:
                per_emp[emp]["breakfast_dine_in"] += cnt
                total_b_dine += cnt
        else:
            if btype == "parcel":
                per_emp[emp]["dinner_parcel"] += cnt
                total_d_parcel += cnt
            else:
                per_emp[emp]["dinner_dine_in"] += cnt
                total_d_dine += cnt
    rows = list(per_emp.values())
    for r in rows:
        r["breakfast"] = r["breakfast_dine_in"] + r["breakfast_parcel"]
        r["dinner"] = r["dinner_dine_in"] + r["dinner_parcel"]
        r["total"] = r["breakfast"] + r["dinner"]
    rows.sort(key=lambda r: r["employee_number"])
    return {
        "from": from_date, "to": to_date,
        "total_breakfast": total_b_dine + total_b_parcel,
        "total_dinner": total_d_dine + total_d_parcel,
        "total_breakfast_dine_in": total_b_dine, "total_breakfast_parcel": total_b_parcel,
        "total_dinner_dine_in": total_d_dine, "total_dinner_parcel": total_d_parcel,
        "employees": rows,
    }


async def build_report_xlsx(from_date: str, to_date: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C65A40")
    headers = [
        "Employee Number", "Name",
        "Breakfast Dine-in", "Breakfast Parcel", "Breakfast Total",
        "Dinner Dine-in", "Dinner Parcel", "Dinner Total",
        "Grand Total",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    pipeline = [
        {"$match": {"meal_date": {"$gte": from_date, "$lte": to_date}, "status": {"$ne": "emergency_cancelled"}}},
        {"$group": {
            "_id": {"emp": "$employee_number", "name": "$employee_name",
                    "meal": "$meal_type", "type": "$booking_type"},
            "qty": {"$sum": "$quantity"},
        }},
    ]
    per_emp: dict = {}
    async for row in db.bookings.aggregate(pipeline):
        emp = row["_id"]["emp"]
        name = row["_id"].get("name") or ""
        meal = row["_id"]["meal"]
        btype = row["_id"].get("type") or "dine_in"
        if emp not in per_emp:
            per_emp[emp] = {
                "emp": emp, "name": name,
                "b_dine": 0, "b_parcel": 0, "d_dine": 0, "d_parcel": 0,
            }
        key = f"{'b' if meal == 'breakfast' else 'd'}_{'parcel' if btype == 'parcel' else 'dine'}"
        per_emp[emp][key] = row["qty"]
    rows = sorted(per_emp.values(), key=lambda r: r["emp"])
    r_idx = 2
    for r in rows:
        b_total = r["b_dine"] + r["b_parcel"]
        d_total = r["d_dine"] + r["d_parcel"]
        grand = b_total + d_total
        ws.cell(row=r_idx, column=1, value=r["emp"])
        ws.cell(row=r_idx, column=2, value=r["name"])
        ws.cell(row=r_idx, column=3, value=r["b_dine"])
        ws.cell(row=r_idx, column=4, value=r["b_parcel"])
        ws.cell(row=r_idx, column=5, value=b_total)
        ws.cell(row=r_idx, column=6, value=r["d_dine"])
        ws.cell(row=r_idx, column=7, value=r["d_parcel"])
        ws.cell(row=r_idx, column=8, value=d_total)
        ws.cell(row=r_idx, column=9, value=grand)
        r_idx += 1
    if rows:
        ws.cell(row=r_idx, column=1, value="TOTAL").font = Font(bold=True)
        totals = {
            3: sum(r["b_dine"] for r in rows),
            4: sum(r["b_parcel"] for r in rows),
            6: sum(r["d_dine"] for r in rows),
            7: sum(r["d_parcel"] for r in rows),
        }
        totals[5] = totals[3] + totals[4]
        totals[8] = totals[6] + totals[7]
        totals[9] = totals[5] + totals[8]
        for col, val in totals.items():
            ws.cell(row=r_idx, column=col, value=val).font = Font(bold=True)
    for col, w in [("A", 20), ("B", 26), ("C", 18), ("D", 18), ("E", 16),
                   ("F", 16), ("G", 16), ("H", 14), ("I", 14)]:
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Bookings")
    headers2 = ["Date", "Meal", "Type", "Qty", "Employee Number", "Name", "Served", "Booked At"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")
    r_idx = 2
    async for b in db.bookings.find({"meal_date": {"$gte": from_date, "$lte": to_date}}).sort([("meal_date", 1), ("meal_type", 1)]):
        ws2.cell(row=r_idx, column=1, value=b["meal_date"])
        ws2.cell(row=r_idx, column=2, value=b["meal_type"].capitalize())
        ws2.cell(row=r_idx, column=3, value=b.get("booking_type", "dine_in").replace("_", " ").title())
        ws2.cell(row=r_idx, column=4, value=b.get("quantity", 1))
        ws2.cell(row=r_idx, column=5, value=b["employee_number"])
        ws2.cell(row=r_idx, column=6, value=b.get("employee_name", ""))
        ws2.cell(row=r_idx, column=7, value="Yes" if b.get("served") else "No")
        ws2.cell(row=r_idx, column=8, value=b.get("created_at", ""))
        r_idx += 1
    for col, w in [("A", 14), ("B", 12), ("C", 12), ("D", 8), ("E", 20), ("F", 26), ("G", 10), ("H", 30)]:
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@api.get("/admin/export")
async def admin_export_excel(
    from_date: str = Query(..., alias="from"),
    to_date: str = Query(..., alias="to"),
    admin: dict = Depends(get_admin_user),
):
    try:
        parse_iso_date(from_date); parse_iso_date(to_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")
    data = await build_report_xlsx(from_date, to_date)
    filename = f"mess_bookings_{from_date}_to_{to_date}.xlsx"
    await audit(admin, "report.export", target=f"{from_date}..{to_date}")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/admin/payroll-export")
async def admin_payroll_export(
    month: Optional[str] = Query(None),
    admin: dict = Depends(get_admin_user),
):
    """Export a ready-to-file payroll ledger for the given month (YYYY-MM).
    Combines each employee's breakfast/dinner counts with current meal prices
    and their ₹ total, ready to hand to finance."""
    if not month:
        n = now_local()
        month = f"{n.year:04d}-{n.month:02d}"
    try:
        parse_iso_date(f"{month}-01")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid month, use YYYY-MM")

    prices = await get_meal_prices()
    pipeline = [
        {"$match": {"meal_date": {"$regex": f"^{month}-"},
                    "status": {"$ne": "emergency_cancelled"}}},
        {"$group": {"_id": {"user_id": "$user_id", "meal_type": "$meal_type"},
                    "qty": {"$sum": "$quantity"}}},
    ]
    counts: dict = {}
    async for row in db.bookings.aggregate(pipeline):
        uid = row["_id"]["user_id"]
        counts.setdefault(uid, {"breakfast": 0, "dinner": 0})
        counts[uid][row["_id"]["meal_type"]] = row["qty"]

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {month}"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="e11d48", end_color="e11d48", fill_type="solid")
    headers = ["Employee ID", "Name", "Role", "Breakfast Meals",
               "Dinner Meals", f"Breakfast ₹ (@ {prices['breakfast']})",
               f"Dinner ₹ (@ {prices['dinner']})", "Total ₹"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    r = 2
    grand_b = grand_d = 0
    grand_b_amt = grand_d_amt = 0.0
    async for u in db.users.find({}).sort("employee_number", 1):
        if u.get("role") == "admin":
            continue  # payroll is for employees & chefs (people who eat)
        uid = str(u["_id"])
        c = counts.get(uid, {"breakfast": 0, "dinner": 0})
        b_amt = round(c["breakfast"] * prices["breakfast"], 2)
        d_amt = round(c["dinner"] * prices["dinner"], 2)
        total = round(b_amt + d_amt, 2)
        ws.cell(row=r, column=1, value=u.get("employee_number", ""))
        ws.cell(row=r, column=2, value=u.get("name", ""))
        ws.cell(row=r, column=3, value=u.get("role", "employee").capitalize())
        ws.cell(row=r, column=4, value=c["breakfast"])
        ws.cell(row=r, column=5, value=c["dinner"])
        ws.cell(row=r, column=6, value=b_amt)
        ws.cell(row=r, column=7, value=d_amt)
        ws.cell(row=r, column=8, value=total).font = Font(bold=True)
        grand_b += c["breakfast"]; grand_d += c["dinner"]
        grand_b_amt += b_amt; grand_d_amt += d_amt
        r += 1

    # Grand totals row
    totals_row = r
    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=totals_row, column=4, value=grand_b).font = Font(bold=True)
    ws.cell(row=totals_row, column=5, value=grand_d).font = Font(bold=True)
    ws.cell(row=totals_row, column=6, value=round(grand_b_amt, 2)).font = Font(bold=True)
    ws.cell(row=totals_row, column=7, value=round(grand_d_amt, 2)).font = Font(bold=True)
    ws.cell(row=totals_row, column=8, value=round(grand_b_amt + grand_d_amt, 2)).font = Font(bold=True)
    for col_letter, w in [("A", 14), ("B", 28), ("C", 12), ("D", 16),
                          ("E", 16), ("F", 18), ("G", 18), ("H", 14)]:
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"payroll_ledger_{month}.xlsx"
    await audit(admin, "payroll.export", target=month)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/reports/download/{token}")
async def download_emailed_report(token: str):
    rec = await db.report_downloads.find_one({"token": token})
    if not rec:
        raise HTTPException(status_code=404, detail="Download link is invalid or expired")
    expires_at = rec.get("expires_at")
    if expires_at is not None:
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=_tz.utc)
        if datetime.now(_tz.utc) > expires_at:
            raise HTTPException(status_code=410, detail="Download link has expired")
    return StreamingResponse(
        io.BytesIO(rec["content"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{rec.get("filename", "report.xlsx")}"'},
    )


@api.get("/admin/audit-logs")
async def admin_audit_logs(limit: int = Query(100, ge=1, le=500), admin: dict = Depends(get_admin_user)):
    out = []
    async for log in db.audit_logs.find({}).sort("timestamp", -1).limit(limit):
        out.append({
            "id": str(log["_id"]),
            "actor_employee_number": log.get("actor_employee_number", ""),
            "actor_role": log.get("actor_role", ""),
            "action": log["action"],
            "target": log.get("target", ""),
            "meta": log.get("meta", {}),
            "timestamp": log.get("timestamp"),
        })
    return out


@api.get("/admin/insights")
async def admin_insights(days: int = Query(14, ge=1, le=90), admin: dict = Depends(get_admin_user)):
    """Daily breakdown + top eaters for the last N days (inclusive of today)."""
    now = now_local().date()
    start = now - timedelta(days=days - 1)
    from_date = start.isoformat()
    to_date = now.isoformat()

    # Daily trend: one row per date + meal_type
    trend_map: dict = {}
    async for row in db.bookings.aggregate([
        {"$match": {"meal_date": {"$gte": from_date, "$lte": to_date}, "status": {"$ne": "emergency_cancelled"}}},
        {"$group": {
            "_id": {"date": "$meal_date", "meal": "$meal_type"},
            "qty": {"$sum": "$quantity"},
        }},
    ]):
        d = row["_id"]["date"]
        m = row["_id"]["meal"]
        if d not in trend_map:
            trend_map[d] = {"date": d, "breakfast": 0, "dinner": 0}
        trend_map[d][m] = row["qty"]

    # Fill missing days with zeros so the chart is dense
    trend = []
    cur = start
    while cur <= now:
        iso = cur.isoformat()
        e = trend_map.get(iso, {"date": iso, "breakfast": 0, "dinner": 0})
        e["total"] = e["breakfast"] + e["dinner"]
        trend.append(e)
        cur = cur + timedelta(days=1)

    # Top eaters over the range
    top_map: dict = {}
    async for row in db.bookings.aggregate([
        {"$match": {"meal_date": {"$gte": from_date, "$lte": to_date}, "status": {"$ne": "emergency_cancelled"}}},
        {"$group": {
            "_id": {"emp": "$employee_number", "name": "$employee_name", "meal": "$meal_type"},
            "qty": {"$sum": "$quantity"},
        }},
    ]):
        emp = row["_id"]["emp"]; name = row["_id"].get("name") or ""; m = row["_id"]["meal"]
        if emp not in top_map:
            top_map[emp] = {"employee_number": emp, "name": name, "breakfast": 0, "dinner": 0}
        top_map[emp][m] = row["qty"]
    top = [{**v, "total": v["breakfast"] + v["dinner"]} for v in top_map.values()]
    top.sort(key=lambda r: r["total"], reverse=True)
    top = top[:10]

    return {"from": from_date, "to": to_date, "days": days, "trend": trend, "top_eaters": top}


@api.post("/admin/employees/bulk")
async def admin_bulk_employees(file: UploadFile = File(...), admin: dict = Depends(get_admin_user)):
    """Bulk create/update employees from an Excel (.xlsx) file.

    Required columns (header row, case-insensitive, spaces/underscores ignored):
      - Employee ID  (alias: employee_number, id)
      - Name         (alias: employee_name, employee name)
      - Password

    On duplicate Employee ID: name and password are UPDATED.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx (Excel) file")
    try:
        from openpyxl import load_workbook  # local import — openpyxl already installed
        raw = await file.read()
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel is empty or missing data rows")

    def norm(v):
        return str(v or "").strip().lower().replace(" ", "").replace("_", "")

    header = [norm(c) for c in rows[0]]
    def find_col(*aliases):
        for a in aliases:
            if a in header:
                return header.index(a)
        return -1

    idx_emp = find_col("employeeid", "employeenumber", "id", "empid")
    idx_name = find_col("name", "employeename", "empname")
    idx_pw = find_col("password", "pwd")
    if idx_emp < 0 or idx_name < 0 or idx_pw < 0:
        raise HTTPException(
            status_code=400,
            detail="Excel must have columns: Employee ID, Name, Password"
        )

    created = 0
    updated = 0
    errors = []
    for i, row in enumerate(rows[1:], start=2):
        emp = str(row[idx_emp] or "").strip()
        name = str(row[idx_name] or "").strip()
        pw = str(row[idx_pw] or "").strip()
        if not emp or not name or not pw:
            if not any([emp, name, pw]):
                continue  # skip fully blank rows
            errors.append({"line": i, "error": "Employee ID, Name and Password are all required"})
            continue
        if len(pw) < 4:
            errors.append({"line": i, "error": "password must be at least 4 characters"})
            continue

        existing = await db.users.find_one({"employee_number": emp})
        if existing:
            await db.users.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "name": name,
                    "password_hash": hash_password(pw),
                    "updated_at": now_local().isoformat(),
                }},
            )
            updated += 1
        else:
            await db.users.insert_one({
                "employee_number": emp,
                "name": name,
                "email": None,
                "password_hash": hash_password(pw),
                "role": "employee",
                "created_at": now_local().isoformat(),
            })
            created += 1

    await audit(admin, "employee.bulk_import",
                meta={"created": created, "updated": updated, "errors": len(errors)})
    return {"created": created, "updated": updated, "errors": errors}


class AdminBookingIn(BaseModel):
    user_id: str
    meal_type: Literal["breakfast", "dinner"]
    meal_date: str
    quantity: int = Field(default=1, ge=1, le=MAX_QTY)
    booking_type: Literal["dine_in", "parcel"] = "dine_in"


class DayCancelIn(BaseModel):
    date: str
    meal_type: Literal["breakfast", "dinner", "both"] = "both"
    reason: str = Field(min_length=1, max_length=500)


class EmergencyCancelIn(BaseModel):
    date: str
    meal_type: Literal["breakfast", "dinner", "both"] = "both"
    reason: str = Field(min_length=1, max_length=500)
    applies_to: Literal["all", "selected"] = "all"
    employee_ids: List[str] = Field(default_factory=list)  # required when applies_to=selected


class AdminBookingOverrideIn(BaseModel):
    user_id: str
    meal_type: Literal["breakfast", "dinner"]
    meal_date: str
    quantity: int = Field(default=1, ge=1, le=MAX_QTY)
    booking_type: Literal["dine_in", "parcel"] = "dine_in"
    reason: str = Field(min_length=1, max_length=500)


@api.get("/admin/bookings")
async def admin_list_bookings(
    date: Optional[str] = None,
    meal_type: Optional[Literal["breakfast", "dinner"]] = None,
    q: Optional[str] = None,
    admin: dict = Depends(get_admin_user),
):
    d = date or now_local().date().isoformat()
    query: dict = {"meal_date": d, "status": {"$ne": "emergency_cancelled"}}
    if meal_type:
        query["meal_type"] = meal_type
    if q:
        query["$or"] = [
            {"employee_number": {"$regex": q, "$options": "i"}},
            {"employee_name": {"$regex": q, "$options": "i"}},
        ]
    items = []
    async for b in db.bookings.find(query).sort([("meal_type", 1), ("employee_number", 1)]):
        items.append({
            "id": str(b["_id"]),
            "user_id": b["user_id"],
            "employee_number": b["employee_number"],
            "employee_name": b.get("employee_name", ""),
            "meal_type": b["meal_type"],
            "meal_date": b["meal_date"],
            "quantity": b.get("quantity", 1),
            "booking_type": b.get("booking_type", "dine_in"),
            "served": b.get("served", False),
            "created_at": b.get("created_at"),
            "admin_override": b.get("admin_override", False),
            "override_reason": b.get("override_reason"),
        })
    return items


@api.post("/admin/bookings")
async def admin_create_booking(body: AdminBookingOverrideIn, background: BackgroundTasks, admin: dict = Depends(get_admin_user)):
    """Force-create a booking for an employee, bypassing cutoff/holiday/emergency checks.
    Requires a mandatory reason. Marks the booking as an admin override and emails the employee."""
    try:
        parse_iso_date(body.meal_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid meal_date, use YYYY-MM-DD")
    try:
        target = await db.users.find_one({"_id": ObjectId(body.user_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user_id")
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    existing = await db.bookings.find_one({
        "user_id": str(target["_id"]),
        "meal_type": body.meal_type,
        "meal_date": body.meal_date,
    })
    if existing:
        await db.bookings.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                "quantity": body.quantity, "booking_type": body.booking_type,
                "status": "active",
                "admin_override": True,
                "override_reason": body.reason,
                "override_by": admin["employee_number"],
                "override_at": now_local().isoformat(),
            }},
        )
        await audit(admin, "admin.booking.override_update", target=str(existing["_id"]),
                    meta={"emp": target["employee_number"], "meal": body.meal_type,
                          "date": body.meal_date, "qty": body.quantity, "reason": body.reason})
        return {"id": str(existing["_id"]), "updated": True}

    doc = {
        "user_id": str(target["_id"]),
        "employee_number": target["employee_number"],
        "employee_name": target.get("name", ""),
        "meal_type": body.meal_type,
        "meal_date": body.meal_date,
        "quantity": body.quantity,
        "booking_type": body.booking_type,
        "status": "active",
        "served": False,
        "served_at": None,
        "served_by": None,
        "created_at": now_local().isoformat(),
        "created_by_admin": admin["employee_number"],
        "admin_override": True,
        "override_reason": body.reason,
        "override_by": admin["employee_number"],
        "override_at": now_local().isoformat(),
    }
    res = await db.bookings.insert_one(doc)
    await audit(admin, "admin.booking.override_create", target=str(res.inserted_id),
                meta={"admin_name": admin.get("name", ""), "employee_number": target["employee_number"],
                      "employee_name": target.get("name", ""), "meal_date": body.meal_date,
                      "meal_type": body.meal_type, "quantity": body.quantity,
                      "booking_type": body.booking_type, "reason": body.reason})

    return {"id": str(res.inserted_id), "created": True}


# ---------------- Emergency Cancellations ----------------
@api.post("/admin/emergency-cancellations")
async def create_emergency_cancellation(body: EmergencyCancelIn, background: BackgroundTasks,
                                         admin: dict = Depends(get_admin_user)):
    try:
        parse_iso_date(body.date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date")
    if body.applies_to == "selected" and not body.employee_ids:
        raise HTTPException(status_code=400, detail="Select at least one employee")

    meal_types = ["breakfast", "dinner"] if body.meal_type == "both" else [body.meal_type]
    bq: dict = {"meal_date": body.date, "meal_type": {"$in": meal_types},
                "status": {"$ne": "emergency_cancelled"}}
    if body.applies_to == "selected":
        bq["user_id"] = {"$in": body.employee_ids}

    affected: list = []
    async for b in db.bookings.find(bq):
        affected.append(b)

    now = now_local().isoformat()
    if affected:
        await db.bookings.update_many(
            {"_id": {"$in": [b["_id"] for b in affected]}},
            {"$set": {"status": "emergency_cancelled",
                      "cancelled_at": now,
                      "cancelled_by": admin["employee_number"]}}
        )
        for b in affected:
            await record_cancellation(b, cancelled_by=admin["employee_number"],
                                       actor_role="admin", reason=f"Emergency: {body.reason}")

    doc = {
        "date": body.date,
        "meal_type": body.meal_type,
        "reason": body.reason.strip(),
        "applies_to": body.applies_to,
        "employee_ids": body.employee_ids if body.applies_to == "selected" else [],
        "active": True,
        "affected_count": len(affected),
        "created_by": admin["employee_number"],
        "created_at": now,
    }
    res = await db.emergency_cancellations.insert_one(doc)

    per_user: dict = {}
    for b in affected:
        per_user.setdefault(b["user_id"], []).append({
            "type": b["meal_type"], "qty": b.get("quantity", 1),
            "booking_type": b.get("booking_type", "dine_in"),
        })
    emailed = 0
    if per_user:
        try:
            docs = db.users.find({"_id": {"$in": [ObjectId(u) for u in per_user.keys()]}})
            async for u in docs:
                meals = per_user[str(u["_id"])]
                if await send_apology_email(u, body.date, meals, body.reason, background):
                    emailed += 1
        except Exception as e:
            logger.warning(f"emergency email dispatch failed: {e}")

    await audit(admin, "admin.emergency.create", target=str(res.inserted_id),
                meta={"date": body.date, "meal": body.meal_type,
                      "affected": len(affected), "emailed": emailed, "reason": body.reason})
    return {"id": str(res.inserted_id), "affected": len(affected), "emailed": emailed}


@api.get("/admin/emergency-cancellations")
async def list_emergency_cancellations(admin: dict = Depends(get_admin_user)):
    out = []
    async for e in db.emergency_cancellations.find({}).sort("created_at", -1):
        out.append({
            "id": str(e["_id"]),
            "date": e["date"],
            "meal_type": e["meal_type"],
            "reason": e.get("reason", ""),
            "applies_to": e.get("applies_to", "all"),
            "employee_ids": e.get("employee_ids", []),
            "active": e.get("active", True),
            "affected_count": e.get("affected_count", 0),
            "created_by": e.get("created_by", ""),
            "created_at": e.get("created_at"),
            "reopened_at": e.get("reopened_at"),
        })
    return out


@api.post("/admin/emergency-cancellations/{ec_id}/reopen")
async def reopen_emergency_cancellation(ec_id: str, admin: dict = Depends(get_admin_user)):
    try:
        oid = ObjectId(ec_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    e = await db.emergency_cancellations.find_one({"_id": oid})
    if not e:
        raise HTTPException(status_code=404, detail="Not found")
    if not e.get("active", True):
        return {"ok": True, "already_reopened": True}
    await db.emergency_cancellations.update_one(
        {"_id": oid},
        {"$set": {"active": False, "reopened_at": now_local().isoformat(),
                  "reopened_by": admin["employee_number"]}},
    )
    await audit(admin, "admin.emergency.reopen", target=ec_id)
    return {"ok": True}


class AdminCancelBookingIn(BaseModel):
    reason: str = Field(default="Cancelled by admin", min_length=1, max_length=500)
    notify: bool = True


@api.delete("/admin/bookings/{booking_id}")
async def admin_cancel_booking(booking_id: str, background: BackgroundTasks,
                                admin: dict = Depends(get_admin_user),
                                reason: str = Query("Cancelled by admin"),
                                notify: bool = Query(True)):
    """Force-cancel any booking. Accepts a reason (query params) and optionally emails the employee."""
    try:
        oid = ObjectId(booking_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    booking = await db.bookings.find_one({"_id": oid})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    reason_txt = (reason or "Cancelled by admin").strip() or "Cancelled by admin"
    await db.bookings.delete_one({"_id": oid})
    await record_cancellation(booking, cancelled_by=admin["employee_number"],
                              actor_role="admin", reason=reason_txt)
    emailed = False
    if notify:
        try:
            target = await db.users.find_one({"_id": ObjectId(booking["user_id"])})
        except Exception:
            target = None
        if target:
            emailed = await send_apology_email(
                target, booking["meal_date"],
                [{"type": booking["meal_type"], "qty": booking.get("quantity", 1),
                  "booking_type": booking.get("booking_type", "dine_in")}],
                reason_txt, background,
            )
    await audit(admin, "admin.booking.cancel", target=booking_id,
                meta={"emp": booking.get("employee_number", ""), "meal": booking["meal_type"],
                      "date": booking["meal_date"], "reason": reason_txt, "emailed": emailed})
    return {"ok": True, "emailed": emailed}


@api.post("/admin/cancel-day")
async def admin_cancel_day(body: DayCancelIn, background: BackgroundTasks, admin: dict = Depends(get_admin_user)):
    """Cancel all bookings for a given date (breakfast, dinner, or both) with a reason,
    and email every affected employee whose account has an email on file."""
    try:
        parse_iso_date(body.date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date")

    query: dict = {"meal_date": body.date}
    if body.meal_type != "both":
        query["meal_type"] = body.meal_type

    affected_bookings = []
    async for b in db.bookings.find(query):
        affected_bookings.append(b)

    # Notify each unique user (aggregate their cancelled meals per user)
    per_user: dict = {}
    for b in affected_bookings:
        uid = b["user_id"]
        per_user.setdefault(uid, {
            "employee_number": b["employee_number"],
            "employee_name": b.get("employee_name", ""),
            "meals": [],
        })
        per_user[uid]["meals"].append({
            "type": b["meal_type"], "qty": b.get("quantity", 1), "booking_type": b.get("booking_type", "dine_in"),
        })
        # Log per-booking cancellation event
        await record_cancellation(b, cancelled_by=admin["employee_number"],
                                  actor_role="admin", reason=body.reason)

    # Delete bookings
    deleted = 0
    if affected_bookings:
        res = await db.bookings.delete_many(query)
        deleted = res.deleted_count

    # Email each affected employee via the shared helper
    emailed = 0
    if per_user:
        try:
            user_docs = db.users.find({"_id": {"$in": [ObjectId(uid) for uid in per_user.keys()]}})
        except Exception:
            user_docs = None
        if user_docs is not None:
            async for u in user_docs:
                info = per_user[str(u["_id"])]
                sent = await send_apology_email(u, body.date, info["meals"], body.reason, background)
                if sent:
                    emailed += 1

    await audit(admin, "admin.day.cancel", target=f"{body.date}:{body.meal_type}",
                meta={"deleted": deleted, "affected_users": len(per_user),
                      "emailed": emailed, "reason": body.reason})
    return {
        "ok": True,
        "deleted": deleted,
        "affected_users": len(per_user),
        "emailed": emailed,
        "affected": [
            {"employee_number": info["employee_number"], "name": info["employee_name"],
             "meals": info["meals"]} for info in per_user.values()
        ],
    }


@api.get("/admin/today")
async def admin_today(admin: dict = Depends(get_admin_user)):
    now = now_local()
    today = now.date().isoformat()
    tomorrow = (now.date() + timedelta(days=1)).isoformat()

    async def sum_qty(match: dict) -> int:
        match_active = {**match, "status": {"$ne": "emergency_cancelled"}}
        cur = db.bookings.aggregate([{"$match": match_active}, {"$group": {"_id": None, "q": {"$sum": "$quantity"}}}])
        async for r in cur:
            return r.get("q", 0)
        return 0

    return {
        "today": today,
        "tomorrow": tomorrow,
        "breakfast_today": await sum_qty({"meal_type": "breakfast", "meal_date": today}),
        "dinner_today": await sum_qty({"meal_type": "dinner", "meal_date": today}),
        "breakfast_tomorrow": await sum_qty({"meal_type": "breakfast", "meal_date": tomorrow}),
        "total_employees": await db.users.count_documents({"role": {"$in": ["employee", "chef"]}}),
    }


@api.get("/")
async def root():
    return {"service": "MessBook API", "ok": True}


    return {"ok": True, "affected_users": len(per_user)}


@api.get("/admin/today")
async def admin_today(admin: dict = Depends(get_admin_user)):
    now = now_local()
    today = now.date().isoformat()
    tomorrow = (now.date() + timedelta(days=1)).isoformat()

    async def sum_qty(match: dict) -> int:
        match_active = {**match, "status": {"$ne": "emergency_cancelled"}}
        cur = db.bookings.aggregate([{"$match": match_active}, {"$group": {"_id": None, "q": {"$sum": "$quantity"}}}])
        async for r in cur:
            return r.get("q", 0)
        return 0

    return {
        "today": today,
        "tomorrow": tomorrow,
        "breakfast_today": await sum_qty({"meal_type": "breakfast", "meal_date": today}),
        "dinner_today": await sum_qty({"meal_type": "dinner", "meal_date": today}),
        "breakfast_tomorrow": await sum_qty({"meal_type": "breakfast", "meal_date": tomorrow}),
        "total_employees": await db.users.count_documents({"role": {"$in": ["employee", "chef"]}}),
    }


@api.get("/")
async def root():
    return {"service": "MessBook API", "ok": True}


# ---------------- Startup ----------------
async def seed_admin():
    emp = os.environ.get("ADMIN_EMPLOYEE_NUMBER", "ADMIN")
    pw = os.environ.get("ADMIN_PASSWORD", "admin@123")
    name = os.environ.get("ADMIN_NAME", "Admin")
    admin_email = os.environ.get("ADMIN_EMAIL") or None
    existing = await db.users.find_one({"employee_number": emp})
    if existing is None:
        await db.users.insert_one({
            "employee_number": emp, "name": name, "email": admin_email,
            "password_hash": hash_password(pw), "role": "admin",
            "created_at": now_local().isoformat(),
        })
        logger.info(f"Seeded admin: {emp}")
    else:
        updates = {}
        if existing.get("role") != "admin":
            updates["role"] = "admin"
        if not verify_password(pw, existing.get("password_hash", "")):
            updates["password_hash"] = hash_password(pw)
        if updates:
            await db.users.update_one({"_id": existing["_id"]}, {"$set": updates})
            logger.info(f"Updated admin: {emp}")


@app.on_event("startup")
async def startup():
    await db.users.create_index("employee_number", unique=True)
    await db.bookings.create_index([("user_id", 1), ("meal_type", 1), ("meal_date", 1)], unique=True)
    await db.bookings.create_index("meal_date")
    await db.holidays.create_index("date", unique=True)
    try:
        await db.sunday_overrides.create_index("date", unique=True)
    except Exception:
        pass
    await db.menus.create_index([("date", 1), ("meal_type", 1)], unique=True)
    await db.weekly_menus.create_index([("day_of_week", 1), ("meal_type", 1)], unique=True)
    await db.audit_logs.create_index("timestamp")
    try:
        await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)
    except Exception:
        pass
    try:
        await db.report_downloads.create_index("expires_at", expireAfterSeconds=0)
    except Exception:
        pass
    await seed_admin()


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

"""
Iteration 3 backend tests:
- Seeded admin 135791 login with admin@123
- POST /api/admin/employees/{id}/send-reset-email
- DELETE /api/admin/bookings/{id}?reason=&notify=  (writes cancellation_events, emails employee)
- DELETE /api/bookings/{id} employee-self-cancel (writes cancellation_events with actor_role='employee')
- POST /api/admin/cancel-day writes one cancellation_events per affected booking
- GET /api/bookings/cancellations?month=YYYY-MM
- GET /api/admin/summary now returns breakfast_dine_in/parcel, dinner_dine_in/parcel + totals
- GET /api/admin/export returns xlsx with new column headers
- POST /api/admin/email-report returns download_link, GET /api/reports/download/{token} streams xlsx
- POST /api/auth/reset-password bug fix regression
"""
import os
import io
import uuid
import asyncio
from datetime import date, timedelta, datetime
from zoneinfo import ZoneInfo

import pytest
import requests
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://meal-reserve-32.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"
TZ = ZoneInfo("Asia/Kolkata")

NEW_ADMIN = "135791"
LEGACY_ADMIN = "626586"
ADMIN_PW = "admin@123"

SUFFIX = str(uuid.uuid4())[:6]
EMP_NUM = f"IT3{SUFFIX}"
EMP_NAME = "Iter3 User"
EMP_PW = "pw12345"
EMP_EMAIL = f"iter3_{SUFFIX}@example.com"

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")


def auth(t):
    return {"Authorization": f"Bearer {t}"}


# ---------- Fixtures ----------
@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{API}/auth/login",
                      json={"employee_number": NEW_ADMIN, "password": ADMIN_PW}, timeout=15)
    assert r.status_code == 200, f"New admin {NEW_ADMIN} login failed: {r.text}"
    d = r.json()
    assert d["user"]["role"] == "admin", f"Expected admin role, got {d['user'].get('role')}"
    return d["token"]


@pytest.fixture(scope="module")
def admin_id(admin_token):
    r = requests.get(f"{API}/auth/me", headers=auth(admin_token), timeout=10)
    return r.json()["id"]


@pytest.fixture(scope="module")
def emp(admin_token):
    """Create a fresh employee with email so we can trigger reset-email/apology emails."""
    r = requests.post(f"{API}/auth/register", json={
        "employee_number": EMP_NUM, "name": EMP_NAME, "password": EMP_PW, "email": EMP_EMAIL,
    }, timeout=15)
    assert r.status_code == 200, f"Register failed: {r.text}"
    d = r.json()
    return {"token": d["token"], "id": d["user"]["id"], "num": EMP_NUM, "email": EMP_EMAIL}


# ---------- Seeded admin ----------
class TestSeededAdmin:
    def test_new_admin_login(self):
        r = requests.post(f"{API}/auth/login",
                          json={"employee_number": NEW_ADMIN, "password": ADMIN_PW}, timeout=10)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["user"]["employee_number"] == NEW_ADMIN
        assert d["user"]["role"] == "admin"

    def test_legacy_admin_still_admin(self):
        r = requests.post(f"{API}/auth/login",
                          json={"employee_number": LEGACY_ADMIN, "password": ADMIN_PW}, timeout=10)
        # Legacy admin may exist with a different pw; just verify it doesn't 500.
        assert r.status_code in (200, 401), r.text
        if r.status_code == 200:
            assert r.json()["user"]["role"] == "admin"


# ---------- send-reset-email ----------
class TestSendResetEmail:
    def test_send_reset_email_ok(self, admin_token, emp):
        r = requests.post(
            f"{API}/admin/employees/{emp['id']}/send-reset-email",
            headers=auth(admin_token), timeout=15)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d.get("ok") is True
        assert d.get("sent_to") == emp["email"]

        # Verify a password_reset_tokens row was created
        async def fetch():
            c = AsyncIOMotorClient(MONGO_URL); dbx = c[DB_NAME]
            rec = await dbx.password_reset_tokens.find_one(
                {"user_id": emp["id"], "used": False}, sort=[("expires_at", -1)])
            c.close()
            return rec
        rec = asyncio.run(fetch())
        assert rec is not None, "Reset token was not persisted"

    def test_send_reset_email_no_email_user_400(self, admin_token):
        """Create an employee with a placeholder email, then blank it in DB, and expect 400."""
        emp_num = f"NOMAIL{SUFFIX}"
        # register (email required by API)
        r = requests.post(f"{API}/auth/register", json={
            "employee_number": emp_num, "name": "NoMail", "password": "pw12345",
            "email": f"nomail_{SUFFIX}@example.com",
        }, timeout=15)
        assert r.status_code == 200
        uid = r.json()["user"]["id"]

        # Blank the email directly in Mongo
        async def clear():
            c = AsyncIOMotorClient(MONGO_URL); dbx = c[DB_NAME]
            from bson import ObjectId as _OID
            await dbx.users.update_one({"_id": _OID(uid)}, {"$set": {"email": None}})
            c.close()
        asyncio.run(clear())

        r2 = requests.post(
            f"{API}/admin/employees/{uid}/send-reset-email",
            headers=auth(admin_token), timeout=15)
        assert r2.status_code == 400, r2.text
        assert "email" in r2.text.lower()

        # cleanup
        requests.delete(f"{API}/admin/employees/{uid}", headers=auth(admin_token), timeout=10)

    def test_send_reset_email_bad_id(self, admin_token):
        r = requests.post(
            f"{API}/admin/employees/xyz-notanid/send-reset-email",
            headers=auth(admin_token), timeout=10)
        assert r.status_code == 400

    def test_send_reset_email_admin_only(self, emp):
        r = requests.post(
            f"{API}/admin/employees/{emp['id']}/send-reset-email",
            headers=auth(emp["token"]), timeout=10)
        assert r.status_code == 403


# ---------- Reset password flow (regression from iter 2 bug) ----------
class TestResetPasswordRegression:
    def test_reset_password_end_to_end(self, admin_token, emp):
        # Trigger forgot
        r = requests.post(f"{API}/auth/forgot-password",
                          json={"employee_number": emp["num"]}, timeout=10)
        assert r.status_code == 200

        async def get_token():
            c = AsyncIOMotorClient(MONGO_URL); dbx = c[DB_NAME]
            rec = await dbx.password_reset_tokens.find_one(
                {"user_id": emp["id"], "used": False}, sort=[("expires_at", -1)])
            c.close()
            return rec
        rec = asyncio.run(get_token())
        assert rec is not None
        token = rec["token"]

        new_pw = "regres1"
        r2 = requests.post(f"{API}/auth/reset-password",
                           json={"token": token, "new_password": new_pw}, timeout=10)
        assert r2.status_code == 200, f"Expected 200, got {r2.status_code}: {r2.text}"

        # Login with new pw
        r3 = requests.post(f"{API}/auth/login",
                           json={"employee_number": emp["num"], "password": new_pw}, timeout=10)
        assert r3.status_code == 200

        # Reset back to original
        new_tok = r3.json()["token"]
        requests.post(f"{API}/auth/change-password", headers=auth(new_tok), json={
            "current_password": new_pw, "new_password": EMP_PW,
        }, timeout=10)


# ---------- Cancellation events & admin cancel with reason/notify ----------
def _within_dinner_cutoff():
    now = datetime.now(tz=TZ)
    return now < now.replace(hour=14, minute=30, second=0, microsecond=0)


def _within_breakfast_cutoff():
    now = datetime.now(tz=TZ)
    return now < now.replace(hour=23, minute=30, second=0, microsecond=0)


class TestCancellationEvents:
    def test_employee_self_cancel_records_event(self, emp):
        if not _within_breakfast_cutoff():
            pytest.skip("Past breakfast cutoff")
        tomorrow = (date.today() + timedelta(days=1)).isoformat()

        # Cleanup pre-existing breakfast booking
        st = requests.get(f"{API}/bookings/status", headers=auth(emp["token"]), timeout=10).json()
        for i in st["items"]:
            if i["meal_type"] == "breakfast" and i["booked"]:
                requests.delete(f"{API}/bookings/{i['booking_id']}",
                                headers=auth(emp["token"]), timeout=10)

        rb = requests.post(f"{API}/bookings", headers=auth(emp["token"]), json={
            "meal_type": "breakfast", "meal_date": tomorrow, "quantity": 2, "booking_type": "parcel",
        }, timeout=10)
        assert rb.status_code == 200, rb.text
        bid = rb.json()["id"]

        # Employee self-cancel
        rc = requests.delete(f"{API}/bookings/{bid}", headers=auth(emp["token"]), timeout=10)
        assert rc.status_code == 200

        # Verify cancellation_events row exists (actor_role='employee')
        month = tomorrow[:7]
        r = requests.get(f"{API}/bookings/cancellations",
                         headers=auth(emp["token"]),
                         params={"month": month}, timeout=10)
        assert r.status_code == 200, r.text
        d = r.json()
        events = d["items"]
        matches = [e for e in events if e["meal_date"] == tomorrow and e["meal_type"] == "breakfast"]
        assert matches, f"No cancellation event found. Got: {events}"
        m = matches[0]
        assert m["actor_role"] == "employee"
        assert m["cancelled_by"] == emp["num"]
        assert m["quantity"] == 2
        assert m["booking_type"] == "parcel"
        assert "cancelled_at" in m

    def test_admin_cancel_with_reason_writes_event_and_notifies(self, admin_token, emp):
        if not _within_breakfast_cutoff():
            pytest.skip("Past breakfast cutoff")
        tomorrow = (date.today() + timedelta(days=1)).isoformat()

        # Ensure no pre-existing booking
        st = requests.get(f"{API}/bookings/status", headers=auth(emp["token"]), timeout=10).json()
        for i in st["items"]:
            if i["meal_type"] == "breakfast" and i["booked"]:
                requests.delete(f"{API}/bookings/{i['booking_id']}",
                                headers=auth(emp["token"]), timeout=10)

        # Employee creates booking
        rb = requests.post(f"{API}/bookings", headers=auth(emp["token"]), json={
            "meal_type": "breakfast", "meal_date": tomorrow, "quantity": 1, "booking_type": "dine_in",
        }, timeout=10)
        assert rb.status_code == 200
        bid = rb.json()["id"]

        # Admin cancels with a reason + notify=true
        reason = "TEST cancelled by admin - gas leak"
        rc = requests.delete(
            f"{API}/admin/bookings/{bid}",
            headers=auth(admin_token),
            params={"reason": reason, "notify": "true"},
            timeout=15,
        )
        assert rc.status_code == 200, rc.text
        d = rc.json()
        assert d["ok"] is True
        assert "emailed" in d

        # Verify cancellation event visible to the affected employee
        r = requests.get(f"{API}/bookings/cancellations",
                         headers=auth(emp["token"]),
                         params={"month": tomorrow[:7]}, timeout=10)
        assert r.status_code == 200
        events = r.json()["items"]
        matches = [e for e in events if e["meal_date"] == tomorrow and e["meal_type"] == "breakfast"
                   and reason in e.get("reason", "")]
        assert matches, f"No admin cancellation event with reason. Got: {events}"
        m = matches[0]
        assert m["actor_role"] == "admin"
        assert m["cancelled_by"] == NEW_ADMIN or m["cancelled_by"] == LEGACY_ADMIN

    def test_admin_cancel_day_records_events(self, admin_token, emp):
        if not _within_breakfast_cutoff():
            pytest.skip("Past breakfast cutoff")
        tomorrow = (date.today() + timedelta(days=1)).isoformat()

        # Cleanup: remove any breakfast holiday for tomorrow
        hlist = requests.get(f"{API}/admin/holidays", headers=auth(admin_token), timeout=10).json()
        for h in hlist:
            if h["date"] == tomorrow:
                requests.delete(f"{API}/admin/holidays/{h['id']}",
                                headers=auth(admin_token), timeout=10)

        # Ensure employee has a fresh booking
        st = requests.get(f"{API}/bookings/status", headers=auth(emp["token"]), timeout=10).json()
        for i in st["items"]:
            if i["meal_type"] == "breakfast" and i["booked"]:
                requests.delete(f"{API}/bookings/{i['booking_id']}",
                                headers=auth(emp["token"]), timeout=10)
        rb = requests.post(f"{API}/bookings", headers=auth(emp["token"]), json={
            "meal_type": "breakfast", "meal_date": tomorrow, "quantity": 1,
        }, timeout=10)
        assert rb.status_code == 200

        reason = "TEST full-day cancel"
        rc = requests.post(f"{API}/admin/cancel-day", headers=auth(admin_token), json={
            "date": tomorrow, "meal_type": "breakfast", "reason": reason,
        }, timeout=15)
        assert rc.status_code == 200, rc.text
        data = rc.json()
        assert data["ok"] is True
        assert data["deleted"] >= 1

        # Cancellation event visible to employee
        r = requests.get(f"{API}/bookings/cancellations",
                         headers=auth(emp["token"]),
                         params={"month": tomorrow[:7]}, timeout=10)
        assert r.status_code == 200
        events = r.json()["items"]
        matches = [e for e in events if e["meal_date"] == tomorrow and reason in e.get("reason", "")]
        assert matches, f"No cancel-day event found. Got: {events}"

    def test_cancellations_month_filter(self, emp):
        # Query far-past month → 0 items
        r = requests.get(f"{API}/bookings/cancellations",
                         headers=auth(emp["token"]),
                         params={"month": "2000-01"}, timeout=10)
        assert r.status_code == 200
        assert r.json()["items"] == []

    def test_cancellations_requires_auth(self):
        r = requests.get(f"{API}/bookings/cancellations", timeout=10)
        assert r.status_code == 401


# ---------- Admin summary/export with parcel/dine-in breakdown ----------
class TestSummaryAndExport:
    def test_summary_returns_new_fields(self, admin_token, emp):
        # Create bookings with different types to test aggregation.
        tomorrow = (date.today() + timedelta(days=1)).isoformat()
        # Cleanup potential existing bookings
        st = requests.get(f"{API}/bookings/status", headers=auth(emp["token"]), timeout=10).json()
        for i in st["items"]:
            if i["meal_type"] == "breakfast" and i["booked"]:
                requests.delete(f"{API}/bookings/{i['booking_id']}", headers=auth(emp["token"]), timeout=10)

        if _within_breakfast_cutoff():
            # breakfast tomorrow parcel qty=2
            r = requests.post(f"{API}/bookings", headers=auth(emp["token"]), json={
                "meal_type": "breakfast", "meal_date": tomorrow, "quantity": 2, "booking_type": "parcel",
            }, timeout=10)
            assert r.status_code == 200

        # Fetch admin summary over the range
        r2 = requests.get(f"{API}/admin/summary", headers=auth(admin_token),
                          params={"from": tomorrow, "to": tomorrow}, timeout=15)
        assert r2.status_code == 200, r2.text
        s = r2.json()
        for f in ("total_breakfast_dine_in", "total_breakfast_parcel",
                  "total_dinner_dine_in", "total_dinner_parcel",
                  "total_breakfast", "total_dinner", "employees"):
            assert f in s, f"missing field {f}"
        # If we created a breakfast parcel qty=2, check aggregation
        if _within_breakfast_cutoff():
            assert s["total_breakfast_parcel"] >= 2, s
            row = next((e for e in s["employees"] if e["employee_number"] == EMP_NUM), None)
            assert row, f"emp row missing: {s['employees']}"
            assert row["breakfast_parcel"] >= 2
            for f in ("breakfast_dine_in", "breakfast_parcel", "dinner_dine_in", "dinner_parcel"):
                assert f in row

    def test_export_xlsx_headers(self, admin_token):
        today = date.today().isoformat()
        week_ago = (date.today() - timedelta(days=7)).isoformat()
        r = requests.get(f"{API}/admin/export", headers=auth(admin_token),
                         params={"from": week_ago, "to": today}, timeout=30)
        assert r.status_code == 200, r.text[:200]
        assert "spreadsheet" in r.headers.get("Content-Type", "").lower() or \
               r.headers.get("Content-Type", "").endswith("sheet")
        wb = load_workbook(io.BytesIO(r.content))
        assert "Summary" in wb.sheetnames
        assert "Bookings" in wb.sheetnames
        ws = wb["Summary"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        expected = ["Employee Number", "Name",
                    "Breakfast Dine-in", "Breakfast Parcel", "Breakfast Total",
                    "Dinner Dine-in", "Dinner Parcel", "Dinner Total",
                    "Grand Total"]
        assert headers == expected, f"Headers mismatch: {headers}"


# ---------- Email report + download link ----------
class TestEmailReportDownload:
    def test_email_report_download_link(self, admin_token):
        today = date.today().isoformat()
        week_ago = (date.today() - timedelta(days=7)).isoformat()
        r = requests.post(f"{API}/admin/email-report", headers=auth(admin_token), json={
            "email": "test@example.com", "from_date": week_ago, "to_date": today,
        }, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["ok"] is True
        assert "download_link" in d and d["download_link"], f"no link in {d}"
        token = d["download_link"].rstrip("/").split("/")[-1]

        # GET the download endpoint (unauthenticated per spec)
        r2 = requests.get(f"{API}/reports/download/{token}", timeout=30)
        assert r2.status_code == 200, r2.text[:200]
        wb = load_workbook(io.BytesIO(r2.content))
        assert "Summary" in wb.sheetnames

    def test_download_bad_token(self):
        r = requests.get(f"{API}/reports/download/does-not-exist", timeout=10)
        assert r.status_code == 404


# ---------- Cleanup ----------
def test_cleanup_iteration3(admin_token):
    emps = requests.get(f"{API}/admin/employees", headers=auth(admin_token), timeout=10).json()
    for e in emps:
        if e["employee_number"].startswith(("IT3", "NOMAIL")):
            requests.delete(f"{API}/admin/employees/{e['id']}",
                            headers=auth(admin_token), timeout=10)

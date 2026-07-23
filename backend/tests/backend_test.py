"""
Backend API tests for Mess Meal Booking System.
Covers auth (register/login), booking CRUD with cutoff, admin employees, admin summary/export.
"""
import os
import io
from datetime import date, timedelta, datetime
from zoneinfo import ZoneInfo

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://meal-reserve-32.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

TZ = ZoneInfo("Asia/Kolkata")

ADMIN_EMP = "626586"
ADMIN_PW = "admin@123"

# Use unique test employees to avoid collisions with prior runs
import uuid
TEST_EMP_SUFFIX = str(uuid.uuid4())[:6]
EMP_NUM = f"TEST{TEST_EMP_SUFFIX}"
EMP_NAME = "Test Employee"
EMP_PW = "test1234"


# -------- Fixtures --------
@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{API}/auth/login", json={"employee_number": ADMIN_EMP, "password": ADMIN_PW}, timeout=15)
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    data = r.json()
    assert data["user"]["role"] == "admin"
    return data["token"]


@pytest.fixture(scope="module")
def emp_token():
    # Register a fresh employee
    r = requests.post(f"{API}/auth/register", json={
        "employee_number": EMP_NUM,
        "name": EMP_NAME,
        "password": EMP_PW,
    }, timeout=15)
    assert r.status_code == 200, f"Register failed: {r.status_code} {r.text}"
    data = r.json()
    assert data["user"]["employee_number"] == EMP_NUM
    assert data["user"]["role"] == "employee"
    return data["token"]


def auth(token):
    return {"Authorization": f"Bearer {token}"}


# -------- Health --------
def test_health_root():
    r = requests.get(f"{API}/", timeout=10)
    assert r.status_code == 200
    assert r.json().get("ok") is True


# -------- Auth --------
def test_admin_login_and_me(admin_token):
    r = requests.get(f"{API}/auth/me", headers=auth(admin_token), timeout=10)
    assert r.status_code == 200
    data = r.json()
    assert data["employee_number"] == ADMIN_EMP
    assert data["role"] == "admin"


def test_invalid_login():
    r = requests.post(f"{API}/auth/login", json={"employee_number": ADMIN_EMP, "password": "wrong"}, timeout=10)
    assert r.status_code == 401


def test_register_duplicate(emp_token):
    r = requests.post(f"{API}/auth/register", json={
        "employee_number": EMP_NUM, "name": "X", "password": "test1234",
    }, timeout=10)
    assert r.status_code == 400


def test_unauth_access():
    r = requests.get(f"{API}/auth/me", timeout=10)
    assert r.status_code == 401
    r2 = requests.get(f"{API}/admin/today", timeout=10)
    assert r2.status_code == 401


def test_non_admin_forbidden(emp_token):
    r = requests.get(f"{API}/admin/today", headers=auth(emp_token), timeout=10)
    assert r.status_code == 403


# -------- Booking status --------
def test_booking_status(emp_token):
    r = requests.get(f"{API}/bookings/status", headers=auth(emp_token), timeout=10)
    assert r.status_code == 200
    data = r.json()
    assert "items" in data
    assert len(data["items"]) == 2
    meal_types = {i["meal_type"] for i in data["items"]}
    assert meal_types == {"breakfast", "dinner"}
    for i in data["items"]:
        assert "cutoff" in i
        assert "cutoff_passed" in i
        assert "booked" in i


# -------- Booking cutoff validation --------
def test_booking_past_date_rejected(emp_token):
    past = (date.today() - timedelta(days=5)).isoformat()
    r = requests.post(f"{API}/bookings", headers=auth(emp_token), json={
        "meal_type": "breakfast", "meal_date": past,
    }, timeout=10)
    assert r.status_code == 400
    assert "cutoff" in r.text.lower() or "past" in r.text.lower()


def test_booking_dinner_today_past_cutoff(emp_token):
    """If it's past 14:30 IST, booking today's dinner should be rejected."""
    now_ist = datetime.now(tz=TZ)
    today = now_ist.date().isoformat()
    r = requests.post(f"{API}/bookings", headers=auth(emp_token), json={
        "meal_type": "dinner", "meal_date": today,
    }, timeout=10)
    cutoff_time = now_ist.replace(hour=14, minute=30, second=0, microsecond=0)
    if now_ist >= cutoff_time:
        assert r.status_code == 400
        assert "cutoff" in r.text.lower()
    else:
        # Should succeed (cleanup)
        assert r.status_code == 200
        bid = r.json()["id"]
        requests.delete(f"{API}/bookings/{bid}", headers=auth(emp_token), timeout=10)


# -------- Booking CRUD --------
def test_book_breakfast_tomorrow_and_cancel(emp_token):
    """Test full flow: create booking for tomorrow's breakfast, verify, then cancel."""
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    # If already booked from prior run, cancel first
    status = requests.get(f"{API}/bookings/status", headers=auth(emp_token), timeout=10).json()
    for item in status["items"]:
        if item["meal_type"] == "breakfast" and item["booked"]:
            requests.delete(f"{API}/bookings/{item['booking_id']}", headers=auth(emp_token), timeout=10)

    # Check breakfast cutoff not passed
    now_ist = datetime.now(tz=TZ)
    tom_date = date.today() + timedelta(days=1)
    cutoff = datetime(tom_date.year, tom_date.month, tom_date.day, 0, 0, tzinfo=TZ) - timedelta(minutes=30)
    # breakfast cutoff = day before at 23:30, so for tomorrow it's today 23:30
    breakfast_cutoff = datetime(now_ist.year, now_ist.month, now_ist.day, 23, 30, tzinfo=TZ)
    if now_ist >= breakfast_cutoff:
        pytest.skip("Past breakfast cutoff for tomorrow, skipping booking test")

    # Create
    r = requests.post(f"{API}/bookings", headers=auth(emp_token), json={
        "meal_type": "breakfast", "meal_date": tomorrow,
    }, timeout=10)
    assert r.status_code == 200, f"Book failed: {r.text}"
    booking = r.json()
    assert booking["meal_type"] == "breakfast"
    assert booking["meal_date"] == tomorrow
    assert "id" in booking
    booking_id = booking["id"]

    # Verify persisted via status
    r2 = requests.get(f"{API}/bookings/status", headers=auth(emp_token), timeout=10)
    items = {i["meal_type"]: i for i in r2.json()["items"]}
    assert items["breakfast"]["booked"] is True
    assert items["breakfast"]["booking_id"] == booking_id

    # Double booking prevention
    r3 = requests.post(f"{API}/bookings", headers=auth(emp_token), json={
        "meal_type": "breakfast", "meal_date": tomorrow,
    }, timeout=10)
    assert r3.status_code == 400
    assert "already" in r3.text.lower()

    # Verify appears in mine
    r4 = requests.get(f"{API}/bookings/mine", headers=auth(emp_token), timeout=10)
    assert r4.status_code == 200
    mine = r4.json()
    assert any(b["id"] == booking_id for b in mine["items"])
    assert mine["breakfast_count"] >= 1

    # Cancel
    r5 = requests.delete(f"{API}/bookings/{booking_id}", headers=auth(emp_token), timeout=10)
    assert r5.status_code == 200
    assert r5.json()["ok"] is True

    # Verify gone
    r6 = requests.get(f"{API}/bookings/status", headers=auth(emp_token), timeout=10)
    items2 = {i["meal_type"]: i for i in r6.json()["items"]}
    assert items2["breakfast"]["booked"] is False


def test_cancel_others_booking_forbidden(emp_token, admin_token):
    """Admin creating a booking, then employee cannot cancel it (403)."""
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    now_ist = datetime.now(tz=TZ)
    breakfast_cutoff = datetime(now_ist.year, now_ist.month, now_ist.day, 23, 30, tzinfo=TZ)
    if now_ist >= breakfast_cutoff:
        pytest.skip("Past cutoff")
    # Clean any existing for admin
    st = requests.get(f"{API}/bookings/status", headers=auth(admin_token), timeout=10).json()
    for item in st["items"]:
        if item["meal_type"] == "breakfast" and item["booked"]:
            requests.delete(f"{API}/bookings/{item['booking_id']}", headers=auth(admin_token), timeout=10)

    r = requests.post(f"{API}/bookings", headers=auth(admin_token), json={
        "meal_type": "breakfast", "meal_date": tomorrow,
    }, timeout=10)
    assert r.status_code == 200
    bid = r.json()["id"]

    # Employee tries to cancel admin's booking
    r2 = requests.delete(f"{API}/bookings/{bid}", headers=auth(emp_token), timeout=10)
    assert r2.status_code == 403

    # Cleanup
    requests.delete(f"{API}/bookings/{bid}", headers=auth(admin_token), timeout=10)


# -------- Admin employee CRUD --------
def test_admin_create_and_delete_employee(admin_token):
    new_emp = f"NEW{TEST_EMP_SUFFIX}"
    r = requests.post(f"{API}/admin/employees", headers=auth(admin_token), json={
        "employee_number": new_emp,
        "name": "Admin Created",
        "password": "pass1234",
        "role": "employee",
    }, timeout=10)
    assert r.status_code == 200, f"Create failed: {r.text}"
    user = r.json()
    assert user["employee_number"] == new_emp
    assert user["role"] == "employee"
    assert "id" in user
    user_id = user["id"]

    # Verify list contains it
    r2 = requests.get(f"{API}/admin/employees", headers=auth(admin_token), timeout=10)
    assert r2.status_code == 200
    emps = r2.json()
    assert any(e["employee_number"] == new_emp for e in emps)

    # Login as new employee
    r3 = requests.post(f"{API}/auth/login", json={"employee_number": new_emp, "password": "pass1234"}, timeout=10)
    assert r3.status_code == 200

    # Delete
    r4 = requests.delete(f"{API}/admin/employees/{user_id}", headers=auth(admin_token), timeout=10)
    assert r4.status_code == 200

    # Verify gone (login fails)
    r5 = requests.post(f"{API}/auth/login", json={"employee_number": new_emp, "password": "pass1234"}, timeout=10)
    assert r5.status_code == 401


def test_admin_cannot_delete_self(admin_token):
    me = requests.get(f"{API}/auth/me", headers=auth(admin_token), timeout=10).json()
    r = requests.delete(f"{API}/admin/employees/{me['id']}", headers=auth(admin_token), timeout=10)
    assert r.status_code == 400


# -------- Admin dashboard --------
def test_admin_today_kpis(admin_token):
    r = requests.get(f"{API}/admin/today", headers=auth(admin_token), timeout=10)
    assert r.status_code == 200
    data = r.json()
    for key in ["today", "tomorrow", "breakfast_today", "dinner_today", "breakfast_tomorrow", "total_employees"]:
        assert key in data
    assert isinstance(data["breakfast_today"], int)
    assert isinstance(data["total_employees"], int)


def test_admin_summary_range(admin_token):
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    r = requests.get(f"{API}/admin/summary", headers=auth(admin_token), params={"from": week_ago, "to": today}, timeout=10)
    assert r.status_code == 200
    data = r.json()
    assert data["from"] == week_ago
    assert data["to"] == today
    assert "employees" in data
    assert "total_breakfast" in data
    assert "total_dinner" in data


def test_admin_summary_invalid_date(admin_token):
    r = requests.get(f"{API}/admin/summary", headers=auth(admin_token), params={"from": "bad", "to": "date"}, timeout=10)
    assert r.status_code == 400


# -------- Excel Export --------
def test_admin_export_excel(admin_token):
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    r = requests.get(f"{API}/admin/export", headers=auth(admin_token), params={"from": week_ago, "to": today}, timeout=30)
    assert r.status_code == 200
    ct = r.headers.get("content-type", "")
    assert "spreadsheetml" in ct or "xlsx" in ct, f"Wrong content type: {ct}"
    cd = r.headers.get("content-disposition", "")
    assert ".xlsx" in cd, f"No xlsx filename in Content-Disposition: {cd}"
    assert f"mess_bookings_{week_ago}_to_{today}.xlsx" in cd
    # Verify openable as xlsx
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert "Summary" in wb.sheetnames
    assert "Bookings" in wb.sheetnames
    ws = wb["Summary"]
    # Header row
    assert ws.cell(row=1, column=1).value == "Employee Number"


def test_export_non_admin_forbidden(emp_token):
    r = requests.get(f"{API}/admin/export", headers=auth(emp_token), params={"from": "2025-01-01", "to": "2025-01-31"}, timeout=15)
    assert r.status_code == 403
